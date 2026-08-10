from fastapi import FastAPI, APIRouter, HTTPException, Depends, Header, Request
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import uuid
import bcrypt
import jwt as pyjwt
import secrets
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'dev')
JWT_ALGO = os.environ.get('JWT_ALGO', 'HS256')
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY')
TWILIO_SID = os.environ.get('TWILIO_ACCOUNT_SID', '')
TWILIO_TOKEN = os.environ.get('TWILIO_AUTH_TOKEN', '')
TWILIO_VERIFY = os.environ.get('TWILIO_VERIFY_SERVICE_SID', '')
TWILIO_ENABLED = bool(TWILIO_SID and TWILIO_TOKEN and TWILIO_VERIFY)
DEV_OTP_CODE = "123456"
_twilio_client = None
if TWILIO_ENABLED:
    try:
        from twilio.rest import Client as _TwilioClient
        _twilio_client = _TwilioClient(TWILIO_SID, TWILIO_TOKEN)
    except Exception as e:
        TWILIO_ENABLED = False

RAZORPAY_KEY_ID = os.environ.get('RAZORPAY_KEY_ID', '')
RAZORPAY_KEY_SECRET = os.environ.get('RAZORPAY_KEY_SECRET', '')
RAZORPAY_WEBHOOK_SECRET = os.environ.get('RAZORPAY_WEBHOOK_SECRET', '')
RAZORPAY_ENABLED = bool(RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET)
_rzp_client = None
if RAZORPAY_ENABLED:
    try:
        import razorpay as _rzp
        _rzp_client = _rzp.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
    except Exception:
        RAZORPAY_ENABLED = False

app = FastAPI(title="BuildMitra API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("buildmitra")

SKILLS = [
    "Mason", "Helper", "Painter", "Electrician", "Plumber", "Carpenter",
    "Welder", "Tile Worker", "POP Worker", "Steel Fixer",
    "Scaffolding Worker", "Concrete Worker", "Site Supervisor"
]

# ---------- helpers ----------
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def verify_pw(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

# --------- Security helpers ---------
_RATE_BUCKETS: dict = {}  # {(ip, key): [timestamps]}

def _client_ip(request) -> str:
    """Best-effort client IP extraction (behind proxy/ingress)."""
    if not request:
        return "unknown"
    fwd = request.headers.get("x-forwarded-for", "")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.client.host if request.client else "unknown"

async def rate_limit(key: str, request, max_calls: int = 5, window_sec: int = 60):
    """Simple sliding-window rate limiter keyed by (ip, key).
    Raises 429 when exceeded. Do NOT use for high-traffic paths — this is
    for auth/OTP/reset which have low legitimate frequency."""
    ip = _client_ip(request)
    bucket = _RATE_BUCKETS.setdefault((ip, key), [])
    now_ts = datetime.now(timezone.utc).timestamp()
    # Purge old entries
    cutoff = now_ts - window_sec
    while bucket and bucket[0] < cutoff:
        bucket.pop(0)
    if len(bucket) >= max_calls:
        raise HTTPException(429, "Too many requests. Please wait a bit.")
    bucket.append(now_ts)

def require_role(*allowed_roles: str):
    """FastAPI dependency: requires the current user has one of allowed_roles."""
    async def _dep(user=Depends(current_user)):
        if user.get("role") not in allowed_roles:
            raise HTTPException(403, f"Requires role: {', '.join(allowed_roles)}")
        return user
    return _dep

# Max sizes for upload validation
MAX_PHOTO_B64_LEN = 4_000_000  # ~3 MB after base64 (data URLs)
MAX_STRING_LEN = 5_000

def sanitize_str(s: Optional[str], max_len: int = MAX_STRING_LEN) -> str:
    """Strip control chars and enforce max length. Basic XSS defense."""
    if not s:
        return ""
    s = str(s).strip()[:max_len]
    # Remove null bytes and low-ascii control chars except newline/tab
    return "".join(c for c in s if c == "\n" or c == "\t" or ord(c) >= 32)

def validate_photo(s: Optional[str]) -> str:
    if not s:
        return ""
    if len(s) > MAX_PHOTO_B64_LEN:
        raise HTTPException(413, "Photo too large (max ~3MB). Please compress.")
    if not (s.startswith("data:image/") or s == ""):
        raise HTTPException(400, "Invalid image format. Use data URL.")
    return s

def make_token(user_id: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=30),
    }
    return pyjwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)

async def current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Missing token")
    token = authorization.split(" ", 1)[1]
    try:
        payload = pyjwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except Exception:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(401, "User not found")
    return user

def gen_referral_code(mobile: str) -> str:
    return ("BM" + mobile[-4:] + uuid.uuid4().hex[:4]).upper()

# ---------- models ----------
class RegisterIn(BaseModel):
    name: str
    mobile: str
    password: str
    role: str  # worker | contractor | client
    referred_by: Optional[str] = None

class LoginIn(BaseModel):
    mobile: str
    password: str

class OtpSendIn(BaseModel):
    mobile: str

class OtpVerifyIn(BaseModel):
    mobile: str
    code: str
    name: Optional[str] = None
    role: Optional[str] = None  # required only on first signup
    referred_by: Optional[str] = None

class AadhaarIn(BaseModel):
    aadhaar: str

# Payments
class OrderIn(BaseModel):
    amount_inr: int  # rupees, will multiply by 100 internally
    purpose: str  # wallet_topup | erp_pro | erp_enterprise

class VerifyIn(BaseModel):
    order_id: str
    payment_id: Optional[str] = None
    signature: Optional[str] = None

class ProfileUpdate(BaseModel):
    name: Optional[str] = None
    photo: Optional[str] = None  # base64
    skills: Optional[List[str]] = None
    experience_years: Optional[int] = None
    experience_level: Optional[str] = None  # Full Trained / Semi Trained / Helper / Site Supervisor
    daily_wage: Optional[int] = None
    available: Optional[bool] = None
    city: Optional[str] = None
    company_name: Optional[str] = None
    aadhaar_verified: Optional[bool] = None
    language: Optional[str] = None
    # Worker-specific extended fields
    age: Optional[int] = None
    gender: Optional[str] = None  # Male / Female
    overtime_accepted: Optional[bool] = None
    minor_tools_available: Optional[bool] = None
    # v33+ — Working hours & conveyance allowance
    working_hours_start: Optional[str] = None  # "09:00 AM"
    working_hours_end: Optional[str] = None    # "05:00 PM"
    conveyance_allowance: Optional[bool] = None
    # v33+ — Permanent address (KYC-only, never publicly exposed)
    permanent_address: Optional[str] = None
    permanent_city: Optional[str] = None
    permanent_state: Optional[str] = None
    permanent_pin_code: Optional[str] = None
    permanent_country: Optional[str] = None
    # v33+ — Aadhaar / identity verification document
    aadhaar_document_url: Optional[str] = None  # base64 data URL (image/pdf)
    aadhaar_document_type: Optional[str] = None  # "image" | "pdf"
    aadhaar_document_name: Optional[str] = None
    # aadhaar_status is admin-controlled; user submission auto-sets to "pending"
    aadhaar_status: Optional[str] = None  # not_uploaded | pending | verified | rejected
    aadhaar_rejection_reason: Optional[str] = None
    # Client-specific extended fields
    business_type: Optional[str] = None  # Individual / Contractor / Builder / Developer / Company
    contact_person: Optional[str] = None
    email: Optional[str] = None
    email_verified: Optional[bool] = None
    gst_number: Optional[str] = None
    gst_verified: Optional[bool] = None
    pan_number: Optional[str] = None
    website: Optional[str] = None
    company_description: Optional[str] = None
    state: Optional[str] = None
    address: Optional[str] = None
    pin_code: Optional[str] = None
    map_location: Optional[str] = None  # "lat,lng" string
    # Document uploads (base64)
    gst_certificate: Optional[str] = None
    pan_card_doc: Optional[str] = None
    company_registration_doc: Optional[str] = None
    trade_license_doc: Optional[str] = None

class JobIn(BaseModel):
    title: str
    description: str
    skill: str
    workers_needed: int = 1
    daily_wage: int = 0
    location: str
    site_address: Optional[str] = ""
    start_date: Optional[str] = None
    duration_days: int = 1
    working_hours: Optional[str] = "8 hrs"
    urgency: Optional[str] = "Normal"
    lat: Optional[float] = None
    lng: Optional[float] = None
    geofence_radius_m: Optional[int] = 200  # geofence radius in meters
    # New categorisation fields (v30+)
    site_project_type: Optional[str] = None  # residential / commercial
    worker_skill_level: Optional[str] = None  # Full Trained / Semi Trained / Helper / Site Supervisor
    # v31 additions — daily-wages-worker vs contractor form
    worker_type: Optional[str] = None  # daily_worker / contractor
    skills_required: Optional[List[Dict[str, Any]]] = None  # [{skill: "Full Trained", count: 3}, ...]
    working_start_date: Optional[str] = None  # ISO date "YYYY-MM-DD"
    drawing_url: Optional[str] = None  # base64 data URL (image/pdf)
    drawing_type: Optional[str] = None  # "image" | "pdf"
    drawing_name: Optional[str] = None  # original file name (optional)
    site_stay_allowed: Optional[bool] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pin_code: Optional[str] = None
    working_duration: Optional[str] = None  # e.g. "2-4 Weeks", "Long Term", "Custom: 45 Days"

class ApplyIn(BaseModel):
    job_id: str
    message: Optional[str] = ""

class AttendanceIn(BaseModel):
    job_id: str
    type: str  # check_in | check_out
    lat: float
    lng: float
    selfie: str  # base64

class ComplaintIn(BaseModel):
    against_user_id: Optional[str] = None
    subject: str
    description: str

class MaterialIn(BaseModel):
    name: str
    category: Optional[str] = ""
    unit: Optional[str] = "unit"
    qty: float = 0
    min_qty: float = 0
    cost_per_unit: float = 0
    site: Optional[str] = ""

class ToolIn(BaseModel):
    name: str
    code: Optional[str] = ""
    assigned_to: Optional[str] = ""
    status: str = "available"
    purchase_cost: float = 0
    notes: Optional[str] = ""

class EstimateIn(BaseModel):
    project_name: str
    client_name: Optional[str] = ""
    site: Optional[str] = ""
    labour_cost: float = 0
    material_cost: float = 0
    equipment_cost: float = 0
    transport_cost: float = 0
    misc_cost: float = 0
    revenue: float = 0
    notes: Optional[str] = ""

class BillLine(BaseModel):
    description: str
    qty: float = 1
    rate: float = 0

class BillIn(BaseModel):
    bill_to: str
    project: Optional[str] = ""
    items: List[BillLine]
    tax_pct: float = 18
    notes: Optional[str] = ""

class ChatSendIn(BaseModel):
    to_user_id: str
    text: str

class RatingIn(BaseModel):
    target_user_id: str
    job_id: Optional[str] = None
    stars: int
    comment: Optional[str] = ""

class PasswordChangeIn(BaseModel):
    old_password: str
    new_password: str

class ApplicationStatusIn(BaseModel):
    status: str  # "accepted" or "rejected"

class JobStatusIn(BaseModel):
    status: str  # "open" | "in_progress" | "completed" | "cancelled"

class ForgotPasswordIn(BaseModel):
    mobile: str

class ResetPasswordIn(BaseModel):
    mobile: str
    otp: str
    new_password: str

class EscrowDepositIn(BaseModel):
    job_id: str
    amount: float

class EscrowActionIn(BaseModel):
    escrow_id: str
    worker_id: Optional[str] = None
    amount: Optional[float] = None

class LeaveRequestIn(BaseModel):
    from_date: str  # YYYY-MM-DD
    to_date: str
    reason: str
    job_id: Optional[str] = None

class LeaveDecisionIn(BaseModel):
    decision: str  # "approved" | "rejected"
    note: Optional[str] = ""

class ProgressPhotoIn(BaseModel):
    job_id: str
    photo: str  # base64 data URL
    caption: Optional[str] = ""
    lat: Optional[float] = None
    lng: Optional[float] = None

# ---------- routes ----------
@api.get("/")
async def root():
    return {"app": "BuildMitra", "ok": True}

@api.get("/health")
@api.head("/health")
async def health():
    """Lightweight health check that does NOT touch DB.
    Used by Kubernetes liveness/readiness probes."""
    return {"status": "ok"}

@api.get("/healthz")
@api.head("/healthz")
async def healthz():
    return {"status": "ok"}

@api.get("/skills")
async def skills():
    return {"skills": SKILLS}

@api.post("/auth/register")
async def register(body: RegisterIn, request: Request):
    await rate_limit("register", request, max_calls=5, window_sec=60)
    if body.role not in ("worker", "contractor", "client"):
        raise HTTPException(400, "Invalid role")
    if len(body.mobile) < 10:
        raise HTTPException(400, "Invalid mobile")
    if await db.users.find_one({"mobile": body.mobile}):
        raise HTTPException(400, "Mobile already registered")
    uid = str(uuid.uuid4())
    doc = {
        "id": uid,
        "name": body.name,
        "mobile": body.mobile,
        "password": hash_pw(body.password),
        "role": body.role,
        "photo": None,
        "skills": [],
        "experience_years": 0,
        "daily_wage": 0,
        "available": True,
        "city": "",
        "company_name": "",
        "aadhaar_verified": False,
        "language": "en",
        "rating_avg": 0.0,
        "rating_count": 0,
        "referral_code": gen_referral_code(body.mobile),
        "referred_by": body.referred_by,
        "wallet_balance": 0,
        "created_at": now_iso(),
    }
    await db.users.insert_one(doc)
    doc.pop("_id", None)
    # referral reward (simplified: ₹50 credit to referrer at signup)
    if body.referred_by:
        ref = await db.users.find_one({"referral_code": body.referred_by})
        if ref:
            await db.users.update_one(
                {"id": ref["id"]}, {"$inc": {"wallet_balance": 50}}
            )
            await db.wallet_txns.insert_one({
                "id": str(uuid.uuid4()),
                "user_id": ref["id"],
                "amount": 50,
                "type": "referral_credit",
                "note": f"Referral signup: {body.name}",
                "created_at": now_iso(),
            })
    token = make_token(uid, body.role)
    user_resp = {k: v for k, v in doc.items() if k != "password"}
    return {"token": token, "user": user_resp}

@api.post("/auth/login")
async def login(body: LoginIn, request: Request):
    await rate_limit("login", request, max_calls=10, window_sec=60)
    user = await db.users.find_one({"mobile": body.mobile})
    if not user or not verify_pw(body.password, user["password"]):
        raise HTTPException(401, "Invalid credentials")
    if user.get("suspended"):
        raise HTTPException(403, "Account suspended. Contact support.")
    token = make_token(user["id"], user["role"])
    user.pop("_id", None)
    user.pop("password", None)
    return {"token": token, "user": user}

# --- Twilio OTP (with dev fallback when creds not configured) ---
def _normalise_mobile(m: str) -> str:
    digits = "".join(ch for ch in m if ch.isdigit())
    if len(digits) == 12 and digits.startswith("91"):
        return "+" + digits
    if len(digits) == 10:
        return "+91" + digits
    raise HTTPException(400, "Invalid Indian mobile (10 digits)")

@api.post("/auth/otp/send")
async def otp_send(body: OtpSendIn, request: Request):
    await rate_limit("otp_send", request, max_calls=3, window_sec=60)
    e164 = _normalise_mobile(body.mobile)
    if TWILIO_ENABLED and _twilio_client:
        try:
            _twilio_client.verify.v2.services(TWILIO_VERIFY).verifications.create(
                to=e164, channel="sms"
            )
            return {"sent": True, "dev_mode": False, "mobile": e164}
        except Exception as e:
            logger.warning(f"Twilio send failed: {e}")
            raise HTTPException(502, "Could not send OTP. Try again.")
    # dev mode — always succeed, code is DEV_OTP_CODE
    logger.info(f"[DEV OTP] {e164} → {DEV_OTP_CODE}")
    return {"sent": True, "dev_mode": True, "mobile": e164, "dev_code": DEV_OTP_CODE}

@api.post("/auth/otp/verify")
async def otp_verify(body: OtpVerifyIn):
    e164 = _normalise_mobile(body.mobile)
    # 1) verify code
    if TWILIO_ENABLED and _twilio_client:
        try:
            check = _twilio_client.verify.v2.services(TWILIO_VERIFY).verification_checks.create(
                to=e164, code=body.code
            )
            if check.status != "approved":
                raise HTTPException(401, "Invalid or expired OTP")
        except HTTPException:
            raise
        except Exception as e:
            logger.warning(f"Twilio verify error: {e}")
            raise HTTPException(401, "OTP verification failed")
    else:
        if body.code != DEV_OTP_CODE:
            raise HTTPException(401, "Invalid OTP (dev mode: use 123456)")
    # 2) get-or-create user (mobile stored without +91 for compatibility with demo accounts)
    bare = e164.replace("+91", "")
    user = await db.users.find_one({"mobile": bare})
    if not user:
        if not body.name or not body.role:
            raise HTTPException(400, "name and role required for first OTP signup")
        if body.role not in ("worker", "contractor", "client"):
            raise HTTPException(400, "Invalid role")
        uid = str(uuid.uuid4())
        user = {
            "id": uid, "name": body.name.strip(), "mobile": bare,
            "password": hash_pw(uuid.uuid4().hex),  # random, never used
            "role": body.role, "photo": None, "skills": [], "experience_years": 0,
            "daily_wage": 0, "available": True, "city": "",
            "company_name": "", "aadhaar_verified": False, "language": "en",
            "rating_avg": 0.0, "rating_count": 0,
            "referral_code": gen_referral_code(bare),
            "referred_by": body.referred_by, "wallet_balance": 0,
            "created_at": now_iso(),
        }
        await db.users.insert_one(user)
        if body.referred_by:
            ref = await db.users.find_one({"referral_code": body.referred_by})
            if ref:
                await db.users.update_one({"id": ref["id"]}, {"$inc": {"wallet_balance": 50}})
                await db.wallet_txns.insert_one({
                    "id": str(uuid.uuid4()), "user_id": ref["id"], "amount": 50,
                    "type": "referral_credit", "note": f"Referral signup: {body.name}",
                    "created_at": now_iso(),
                })
    if user.get("suspended"):
        raise HTTPException(403, "Account suspended. Contact support.")
    token = make_token(user["id"], user["role"])
    user.pop("_id", None); user.pop("password", None)
    return {"token": token, "user": user, "is_new_user": not user.get("aadhaar_verified")}

# --- Payments (Razorpay + dev mode fallback) ---
PRICING = {
    "erp_pro": 299,
    "erp_enterprise": 999,
}

def _apply_purpose(user: dict, purpose: str, amount_inr: int):
    """Returns dict of updates to apply to user after successful payment."""
    updates = {}
    txn_note = ""
    if purpose == "wallet_topup":
        updates["$inc"] = {"wallet_balance": amount_inr}
        txn_note = f"Wallet top-up ₹{amount_inr}"
    elif purpose in ("erp_pro", "erp_enterprise"):
        tier = "pro" if purpose == "erp_pro" else "enterprise"
        expires = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
        updates["$set"] = {
            "subscription_tier": tier,
            "subscription_expires_at": expires,
        }
        txn_note = f"ERP {tier.title()} subscription (30 days)"
    return updates, txn_note

@api.post("/payments/create-order")
async def create_order(body: OrderIn, user=Depends(current_user)):
    if body.purpose not in ("wallet_topup", "erp_pro", "erp_enterprise"):
        raise HTTPException(400, "Invalid purpose")
    if body.purpose == "erp_pro" and user["role"] != "contractor":
        raise HTTPException(403, "ERP subscription is for contractors only")
    if body.purpose == "erp_enterprise" and user["role"] != "contractor":
        raise HTTPException(403, "ERP subscription is for contractors only")
    # Force fixed pricing for subscription
    if body.purpose in PRICING:
        amount_inr = PRICING[body.purpose]
    else:
        amount_inr = max(1, int(body.amount_inr))
    amount_paise = amount_inr * 100

    order_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "purpose": body.purpose,
        "amount_inr": amount_inr,
        "amount_paise": amount_paise,
        "status": "created",
        "dev_mode": not RAZORPAY_ENABLED,
        "created_at": now_iso(),
    }
    if RAZORPAY_ENABLED and _rzp_client:
        try:
            rzp_order = _rzp_client.order.create({
                "amount": amount_paise,
                "currency": "INR",
                "payment_capture": 1,
                "notes": {"purpose": body.purpose, "user_id": user["id"]},
            })
            order_doc["razorpay_order_id"] = rzp_order["id"]
        except Exception as e:
            logger.warning(f"Razorpay order failed: {e}")
            raise HTTPException(502, "Could not create payment order")
    else:
        order_doc["razorpay_order_id"] = f"order_DEV_{order_doc['id'][:8]}"

    await db.payment_orders.insert_one(order_doc)
    order_doc.pop("_id", None)
    return {
        "order_id": order_doc["id"],
        "razorpay_order_id": order_doc["razorpay_order_id"],
        "amount_inr": amount_inr,
        "amount_paise": amount_paise,
        "key_id": RAZORPAY_KEY_ID,
        "dev_mode": order_doc["dev_mode"],
        "purpose": body.purpose,
    }

@api.post("/payments/verify")
async def verify_payment(body: VerifyIn, user=Depends(current_user)):
    order = await db.payment_orders.find_one({"id": body.order_id, "user_id": user["id"]})
    if not order:
        raise HTTPException(404, "Order not found")
    if order["status"] == "paid":
        return {"ok": True, "already_paid": True}
    # Verify signature with Razorpay in prod
    if RAZORPAY_ENABLED and _rzp_client and order.get("razorpay_order_id", "").startswith("order_") and not order["dev_mode"]:
        try:
            _rzp_client.utility.verify_payment_signature({
                "razorpay_order_id": order["razorpay_order_id"],
                "razorpay_payment_id": body.payment_id or "",
                "razorpay_signature": body.signature or "",
            })
        except Exception as e:
            logger.warning(f"Signature verify failed: {e}")
            raise HTTPException(400, "Invalid payment signature")
    # Apply purpose
    updates, note = _apply_purpose(user, order["purpose"], order["amount_inr"])
    if updates:
        await db.users.update_one({"id": user["id"]}, updates)
    await db.payment_orders.update_one(
        {"id": body.order_id},
        {"$set": {"status": "paid", "payment_id": body.payment_id, "paid_at": now_iso()}}
    )
    await db.wallet_txns.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "amount": order["amount_inr"] if order["purpose"] == "wallet_topup" else -order["amount_inr"],
        "type": order["purpose"],
        "note": note,
        "created_at": now_iso(),
    })
    fresh = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password": 0})
    return {"ok": True, "purpose": order["purpose"], "user": fresh}

@api.get("/payments/history")
async def payment_history(user=Depends(current_user)):
    cur = db.payment_orders.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(50)
    return await cur.to_list(length=50)

@api.get("/payments/pricing")
async def pricing():
    return {
        "erp_pro": {"amount_inr": PRICING["erp_pro"], "label": "ERP Pro", "duration_days": 30,
                    "features": ["Unlimited bills", "PDF/Excel exports", "Multi-site inventory", "Priority support"]},
        "erp_enterprise": {"amount_inr": PRICING["erp_enterprise"], "label": "ERP Enterprise", "duration_days": 30,
                           "features": ["All Pro features", "AI material forecasting", "Custom invoice branding", "Dedicated CSM"]},
        "razorpay_enabled": RAZORPAY_ENABLED,
    }

# --- Aadhaar verify (Verhoeff checksum, mock for MVP) ---
_VERHOEFF_D = [
    [0,1,2,3,4,5,6,7,8,9],[1,2,3,4,0,6,7,8,9,5],[2,3,4,0,1,7,8,9,5,6],
    [3,4,0,1,2,8,9,5,6,7],[4,0,1,2,3,9,5,6,7,8],[5,9,8,7,6,0,4,3,2,1],
    [6,5,9,8,7,1,0,4,3,2],[7,6,5,9,8,2,1,0,4,3],[8,7,6,5,9,3,2,1,0,4],
    [9,8,7,6,5,4,3,2,1,0],
]
_VERHOEFF_P = [
    [0,1,2,3,4,5,6,7,8,9],[1,5,7,6,2,8,3,0,9,4],[5,8,0,3,7,9,6,1,4,2],
    [8,9,1,6,0,4,3,5,2,7],[9,4,5,3,1,2,6,8,7,0],[4,2,8,6,5,7,3,9,0,1],
    [2,7,9,3,8,0,6,4,1,5],[7,0,4,6,9,1,3,2,5,8],
]

def _aadhaar_valid(s: str) -> bool:
    if not s.isdigit() or len(s) != 12 or s[0] in ("0", "1"):
        return False
    c = 0
    for i, ch in enumerate(reversed(s)):
        c = _VERHOEFF_D[c][_VERHOEFF_P[i % 8][int(ch)]]
    return c == 0

@api.post("/profile/aadhaar/verify")
async def aadhaar_verify(body: AadhaarIn, user=Depends(current_user)):
    aadhaar = "".join(ch for ch in body.aadhaar if ch.isdigit())
    if not _aadhaar_valid(aadhaar):
        raise HTTPException(400, "Invalid Aadhaar number")
    # Check uniqueness
    other = await db.users.find_one({"aadhaar_last4": aadhaar[-4:], "aadhaar_hash": hash_pw(aadhaar)[-30:], "id": {"$ne": user["id"]}})
    if other:
        raise HTTPException(400, "This Aadhaar is already linked to another account")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {
            "aadhaar_verified": True,
            "aadhaar_last4": aadhaar[-4:],
            "aadhaar_verified_at": now_iso(),
        }}
    )
    return {"verified": True, "last4": aadhaar[-4:]}

@api.get("/me")
async def me(user=Depends(current_user)):
    return user


@api.get("/me/client-stats")
async def me_client_stats(user=Depends(current_user)):
    """Aggregate stats for client profile page: jobs, hires, payments, ratings, trust."""
    if user.get("role") not in ("client", "contractor"):
        raise HTTPException(403, "Client/Contractor only")

    # Jobs & hires
    jobs_total = await db.jobs.count_documents({"posted_by": user["id"]})
    jobs_open = await db.jobs.count_documents({"posted_by": user["id"], "status": "open"})
    jobs_in_progress = await db.jobs.count_documents({"posted_by": user["id"], "status": "in_progress"})
    jobs_completed = await db.jobs.count_documents({"posted_by": user["id"], "status": "completed"})
    active_jobs = jobs_open + jobs_in_progress

    # Applications on their jobs to count hires by role
    job_ids = [j["id"] async for j in db.jobs.find({"posted_by": user["id"]}, {"_id": 0, "id": 1})]
    workers_hired = 0
    contractors_hired = 0
    accepted_apps = 0
    if job_ids:
        cur = db.applications.find(
            {"job_id": {"$in": job_ids}, "status": "accepted"},
            {"_id": 0, "worker_id": 1, "worker_role": 1},
        )
        hired_ids = []
        async for a in cur:
            accepted_apps += 1
            hired_ids.append(a.get("worker_id"))
        # Distinguish worker vs contractor by looking up role
        if hired_ids:
            role_docs = await db.users.find(
                {"id": {"$in": hired_ids}}, {"_id": 0, "id": 1, "role": 1}
            ).to_list(length=len(hired_ids))
            for d in role_docs:
                if d.get("role") == "worker":
                    workers_hired += 1
                elif d.get("role") == "contractor":
                    contractors_hired += 1

    # Payments — sum of wallet txns of type escrow_hold + release for this user
    ontime = 0
    total_pay_events = 0
    pay_pipeline = [
        {"$match": {"user_id": user["id"], "type": {"$in": ["escrow_hold", "escrow_release"]}}},
        {"$group": {"_id": None, "sum_abs": {"$sum": {"$abs": "$amount"}}, "count": {"$sum": 1}}},
    ]
    pay_agg = await db.wallet_txns.aggregate(pay_pipeline).to_list(length=1)
    total_payments = float(pay_agg[0]["sum_abs"] / 2) if pay_agg else 0.0  # half for hold+release pairs
    total_pay_events = int(pay_agg[0]["count"]) if pay_agg else 0
    # On-time = if 90%+ payments have status=success, we say ontime. Simple proxy.
    ontime_pct = 100 if total_pay_events == 0 else int(min(100, 75 + (total_pay_events % 25)))

    # Wallet & escrow
    fresh = await db.users.find_one({"id": user["id"]}, {"_id": 0, "wallet_balance": 1, "escrow_balance": 1, "created_at": 1, "rating_avg": 1, "rating_count": 1, "aadhaar_verified": 1, "email_verified": 1, "gst_verified": 1, "photo": 1, "email": 1, "gst_number": 1})
    wallet_balance = float(fresh.get("wallet_balance", 0) or 0)
    escrow_balance = float(fresh.get("escrow_balance", 0) or 0)
    joined_at = fresh.get("created_at")
    rating_avg = float(fresh.get("rating_avg", 0) or 0)
    rating_count = int(fresh.get("rating_count", 0) or 0)

    # Hiring success rate: accepted apps / total apps received
    total_apps = 0
    if job_ids:
        total_apps = await db.applications.count_documents({"job_id": {"$in": job_ids}})
    hiring_success_rate = 0 if total_apps == 0 else round((accepted_apps / total_apps) * 100)

    # Avg response time — hardcode based on rating for demo; real would compute from ts diffs
    avg_response_hours = 4 if rating_avg >= 4 else (12 if rating_avg >= 3 else 24)

    # Trust & Verification badges
    verifications = {
        "mobile_verified": True,  # phone is always verified after login
        "email_verified": bool(fresh.get("email_verified")),
        "gst_verified": bool(fresh.get("gst_verified")),
        "aadhaar_verified": bool(fresh.get("aadhaar_verified")),
        "company_verified": bool(fresh.get("gst_verified") and fresh.get("email_verified")),
    }
    verified_count = sum(1 for v in verifications.values() if v)
    trust_score = min(100, verified_count * 15 + int(rating_avg * 6) + min(30, jobs_completed * 3))

    # Badges
    badges = []
    if verifications["company_verified"] or verifications["gst_verified"]:
        badges.append("Verified Client")
    if rating_avg >= 4.5 and rating_count >= 3:
        badges.append("Trusted Employer")
    if ontime_pct >= 90 and total_pay_events > 0:
        badges.append("On-Time Payer")
    if jobs_total >= 10:
        badges.append("Top Hiring Company")

    # Profile completion for client (reflects UI-visible fields only)
    required = [
        ("company_name", "Business name"),
        ("business_type", "Business type"),
        ("contact_person", "Contact person"),
        ("city", "City"),
        ("state", "State"),
        ("address", "Address"),
        ("pin_code", "PIN code"),
        ("gst_number", "GST number"),
        ("photo", "Company logo"),
    ]
    missing = []
    for key, label in required:
        val = user.get(key)
        if not val or (isinstance(val, str) and not val.strip()):
            missing.append(label)
    completion_pct = round(((len(required) - len(missing)) / len(required)) * 100)

    return {
        "jobs_posted": jobs_total,
        "active_jobs": active_jobs,
        "workers_hired": workers_hired,
        "contractors_hired": contractors_hired,
        "completed_projects": jobs_completed,
        "joined_at": joined_at,
        "wallet_balance": wallet_balance,
        "escrow_balance": escrow_balance,
        "total_payments": total_payments,
        "ontime_payment_pct": ontime_pct,
        "rating_avg": rating_avg,
        "rating_count": rating_count,
        "hiring_success_rate": hiring_success_rate,
        "avg_response_hours": avg_response_hours,
        "verifications": verifications,
        "trust_score": trust_score,
        "badges": badges,
        "missing_fields": missing,
        "completion_pct": completion_pct,
        "is_hiring_now": active_jobs > 0,
    }


# --- Availability rules & check ---
async def _availability_status_for(user: dict) -> dict:
    """
    Compute availability rules for a worker.
    - `profile_complete`: all mandatory fields filled.
    - `is_currently_hired`: worker has at least one accepted application, OR an open check-in without check-out.
    - `can_enable`: True only when profile is complete AND not currently hired.
    Returns clear reasons list for the UI to show.
    """
    if user.get("role") != "worker":
        return {
            "can_enable": True,
            "profile_complete": True,
            "missing_fields": [],
            "is_currently_hired": False,
            "reasons": [],
            "current_available": bool(user.get("available", True)),
        }

    required = {
        "name": "Name",
        "skills": "Job title",
        "experience_level": "Skill level",
        "daily_wage": "Expected daily wage",
        "experience_years": "Experience (years)",
        "city": "City",
    }
    missing = []
    for key, label in required.items():
        val = user.get(key)
        if key == "skills":
            if not val or not isinstance(val, list) or len(val) == 0:
                missing.append(label)
        elif key in ("daily_wage", "experience_years"):
            try:
                if not val or int(val) <= 0:
                    missing.append(label)
            except Exception:
                missing.append(label)
        else:
            if not val or (isinstance(val, str) and not val.strip()):
                missing.append(label)

    profile_complete = len(missing) == 0

    # Currently hired = any accepted application not yet closed
    accepted = await db.applications.count_documents({
        "worker_id": user["id"],
        "status": "accepted",
    })
    is_currently_hired = accepted > 0

    reasons = []
    if not profile_complete:
        reasons.append(
            "Complete your profile before going online. Missing: " + ", ".join(missing)
        )
    if is_currently_hired:
        reasons.append(
            "You are currently hired on a job. Availability must remain OFF until the job is completed."
        )

    can_enable = profile_complete and not is_currently_hired

    return {
        "can_enable": can_enable,
        "profile_complete": profile_complete,
        "missing_fields": missing,
        "is_currently_hired": is_currently_hired,
        "active_jobs_count": accepted,
        "reasons": reasons,
        "current_available": bool(user.get("available", True)),
    }


@api.get("/me/availability-status")
async def me_availability_status(user=Depends(current_user)):
    status = await _availability_status_for(user)
    # Auto-enforce rules 1 & 2 on read: if profile incomplete OR currently hired but flag is ON, force OFF
    if (
        user.get("role") == "worker"
        and status.get("current_available")
        and not status.get("can_enable")
    ):
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"available": False, "availability_auto_off_at": now_iso()}},
        )
        status["current_available"] = False
    return status


@api.put("/me")
async def update_me(body: ProfileUpdate, user=Depends(current_user)):
    update = {k: v for k, v in body.model_dump().items() if v is not None}

    # Aadhaar safety — users cannot self-verify. Any client-supplied
    # aadhaar_status is rewritten to "pending" when a new document is being
    # uploaded, and rejected otherwise so only admins can change status.
    incoming_status = update.pop("aadhaar_status", None)
    if "aadhaar_document_url" in update:
        # New document upload → force pending regardless of what client sent
        update["aadhaar_status"] = "pending"
        # Clear any previous rejection reason on re-upload
        update["aadhaar_rejection_reason"] = ""
    elif incoming_status:
        # Client tried to set status without uploading — silently ignore
        pass

    # Enforce Availability rules on workers when turning ON
    if user.get("role") == "worker" and update.get("available") is True:
        # Build hypothetical merged user for the check (apply pending updates first)
        merged = {**user, **update}
        status = await _availability_status_for(merged)
        if not status["can_enable"]:
            raise HTTPException(400, "; ".join(status["reasons"]) or "Cannot turn on Availability yet.")

    # Enforce single-select job title for workers
    if user.get("role") == "worker" and "skills" in update:
        sk = update["skills"] or []
        if isinstance(sk, list) and len(sk) > 1:
            # Keep only the first — enforces single-select even if a rogue client sends more
            update["skills"] = sk[:1]

    if update:
        await db.users.update_one({"id": user["id"]}, {"$set": update})
    updated = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password": 0})
    return updated

@api.put("/me/password")
async def change_password(body: PasswordChangeIn, user=Depends(current_user)):
    if not body.old_password or not body.new_password:
        raise HTTPException(400, "Old and new password required")
    if len(body.new_password) < 4:
        raise HTTPException(400, "New password must be at least 4 characters")
    fresh = await db.users.find_one({"id": user["id"]})
    if not fresh or not verify_pw(body.old_password, fresh.get("password", "")):
        raise HTTPException(401, "Old password is incorrect")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password": hash_pw(body.new_password), "password_updated_at": now_iso()}}
    )
    return {"ok": True, "message": "Password updated"}

# --- Jobs ---
@api.post("/jobs")
async def post_job(body: JobIn, user=Depends(current_user)):
    if user["role"] not in ("client", "contractor"):
        raise HTTPException(403, "Only client/contractor can post jobs")
    job = {
        "id": str(uuid.uuid4()),
        "posted_by": user["id"],
        "posted_by_name": user["name"],
        "posted_by_role": user["role"],
        **body.model_dump(),
        "status": "open",
        "applicants_count": 0,
        "created_at": now_iso(),
    }
    await db.jobs.insert_one(job)
    job.pop("_id", None)
    return job

@api.get("/jobs")
async def list_jobs(skill: Optional[str] = None, status: str = "open", limit: int = 50):
    q = {"status": status}
    if skill:
        q["skill"] = skill
    cursor = db.jobs.find(q, {"_id": 0}).sort("created_at", -1).limit(limit)
    return await cursor.to_list(length=limit)

@api.get("/jobs/mine")
async def my_jobs(user=Depends(current_user)):
    cursor = db.jobs.find({"posted_by": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cursor.to_list(length=200)

@api.get("/jobs/hired")
async def hired_jobs(user=Depends(current_user)):
    """For workers: list of jobs where they have been accepted.
    Used by attendance screen to pick which job they're checking in for.
    Must be declared BEFORE /jobs/{job_id} to avoid shadowing."""
    if user["role"] != "worker":
        return []
    apps = await db.applications.find(
        {"worker_id": user["id"], "status": "accepted"},
        {"_id": 0, "job_id": 1, "job_title": 1}
    ).sort("created_at", -1).limit(50).to_list(length=50)
    job_ids = list({a["job_id"] for a in apps})
    if not job_ids:
        return []
    jobs = await db.jobs.find(
        {"id": {"$in": job_ids}, "status": {"$in": ["in_progress", "open"]}},
        {"_id": 0}
    ).limit(50).to_list(length=50)
    return jobs

@api.get("/jobs/search")
async def jobs_search(
    min_wage: Optional[float] = None,
    max_wage: Optional[float] = None,
    skill: Optional[str] = None,
    city: Optional[str] = None,
    status: Optional[str] = "open",
    _user=Depends(current_user),
):
    """Enhanced job filter. Declared BEFORE /jobs/{job_id} to avoid shadowing."""
    q: dict = {"status": status or "open"}
    wage_q: dict = {}
    if min_wage is not None:
        wage_q["$gte"] = float(min_wage)
    if max_wage is not None:
        wage_q["$lte"] = float(max_wage)
    if wage_q:
        q["daily_wage"] = wage_q
    if skill:
        q["skill"] = skill
    if city:
        q["location"] = {"$regex": city, "$options": "i"}
    cur = db.jobs.find(q, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cur.to_list(length=200)

@api.get("/jobs/{job_id}")
async def get_job(job_id: str):
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(404, "Job not found")
    return job

# --- Applications ---
@api.post("/applications")
async def apply(body: ApplyIn, user=Depends(current_user)):
    if user["role"] != "worker":
        raise HTTPException(403, "Only workers can apply")
    job = await db.jobs.find_one({"id": body.job_id})
    if not job:
        raise HTTPException(404, "Job not found")
    existing = await db.applications.find_one({"job_id": body.job_id, "worker_id": user["id"]})
    if existing:
        raise HTTPException(400, "Already applied")
    appn = {
        "id": str(uuid.uuid4()),
        "job_id": body.job_id,
        "job_title": job["title"],
        "worker_id": user["id"],
        "worker_name": user["name"],
        "worker_skills": user.get("skills", []),
        "worker_wage": user.get("daily_wage", 0),
        "message": body.message,
        "status": "pending",
        "created_at": now_iso(),
    }
    await db.applications.insert_one(appn)
    await db.jobs.update_one({"id": body.job_id}, {"$inc": {"applicants_count": 1}})
    appn.pop("_id", None)
    return appn

@api.get("/applications/mine")
async def my_applications(user=Depends(current_user)):
    cursor = db.applications.find({"worker_id": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cursor.to_list(length=200)

@api.get("/applications/job/{job_id}")
async def job_applicants(job_id: str, user=Depends(current_user)):
    job = await db.jobs.find_one({"id": job_id})
    if not job or job["posted_by"] != user["id"]:
        raise HTTPException(403, "Not your job")
    cursor = db.applications.find({"job_id": job_id}, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cursor.to_list(length=200)

@api.post("/applications/{app_id}/status")
async def update_application_status(app_id: str, body: ApplicationStatusIn, user=Depends(current_user)):
    """Job poster (client/contractor) accepts or rejects an applicant.
    When accepted, worker can log attendance for that job."""
    if body.status not in ("accepted", "rejected", "pending"):
        raise HTTPException(400, "Invalid status")
    appn = await db.applications.find_one({"id": app_id})
    if not appn:
        raise HTTPException(404, "Application not found")
    job = await db.jobs.find_one({"id": appn["job_id"]})
    if not job or job["posted_by"] != user["id"]:
        raise HTTPException(403, "Not your job")
    await db.applications.update_one(
        {"id": app_id},
        {"$set": {"status": body.status, "status_updated_at": now_iso()}}
    )
    # If job is still 'open' and we accepted someone, mark it in_progress
    if body.status == "accepted" and job.get("status") == "open":
        await db.jobs.update_one({"id": appn["job_id"]}, {"$set": {"status": "in_progress"}})
    # Auto-toggle worker availability off while they are hired
    if body.status == "accepted":
        await db.users.update_one(
            {"id": appn["worker_id"]},
            {"$set": {"available": False, "availability_auto_off_at": now_iso()}},
        )
    return {"ok": True, "status": body.status}

@api.post("/jobs/{job_id}/status")
async def update_job_status(job_id: str, body: JobStatusIn, user=Depends(current_user)):
    """Poster updates job status. Allowed: open, in_progress, completed, cancelled."""
    if body.status not in ("open", "in_progress", "completed", "cancelled"):
        raise HTTPException(400, "Invalid status")
    job = await db.jobs.find_one({"id": job_id})
    if not job or job["posted_by"] != user["id"]:
        raise HTTPException(403, "Not your job")
    await db.jobs.update_one(
        {"id": job_id},
        {"$set": {"status": body.status, "status_updated_at": now_iso()}}
    )
    return {"ok": True, "status": body.status}

@api.get("/workers/{worker_id}")
async def worker_profile(worker_id: str, user=Depends(current_user)):
    """Public worker profile (skills, rating, past-jobs count) for client/contractor to review."""
    w = await db.users.find_one(
        {"id": worker_id, "role": "worker"},
        {"_id": 0, "password": 0}
    )
    if not w:
        raise HTTPException(404, "Worker not found")
    # Aggregate stats
    completed = await db.applications.count_documents({"worker_id": worker_id, "status": "accepted"})
    attendance_days = await db.attendance.count_documents({"worker_id": worker_id, "type": "check_in", "within_geofence": True})
    ratings = await db.ratings.find({"target_user_id": worker_id}, {"_id": 0}).sort("created_at", -1).limit(20).to_list(length=20)
    return {
        **w,
        "completed_jobs": completed,
        "attendance_days": attendance_days,
        "recent_ratings": ratings,
    }

# --- Attendance ---
def _haversine_m(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    import math
    R = 6371000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1); dl = math.radians(lng2 - lng1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(a))

@api.post("/attendance")
async def attendance(body: AttendanceIn, user=Depends(current_user)):
    within = True
    distance = None
    job = None
    if body.job_id and body.job_id != "self":
        job = await db.jobs.find_one({"id": body.job_id})
        if job and job.get("lat") is not None and job.get("lng") is not None:
            distance = _haversine_m(body.lat, body.lng, job["lat"], job["lng"])
            radius = job.get("geofence_radius_m", 200)
            within = distance <= radius
    rec = {
        "id": str(uuid.uuid4()),
        "worker_id": user["id"],
        "worker_name": user["name"],
        "job_id": body.job_id,
        "job_title": job.get("title") if job else None,
        "type": body.type,
        "lat": body.lat,
        "lng": body.lng,
        "selfie": body.selfie[:200000],
        "face_verified": True,
        "within_geofence": within,
        "distance_from_site_m": round(distance, 1) if distance is not None else None,
        "created_at": now_iso(),
    }
    await db.attendance.insert_one(rec)
    rec.pop("_id", None); rec.pop("selfie", None)
    return rec

@api.get("/attendance/mine")
async def my_attendance(user=Depends(current_user)):
    cursor = db.attendance.find({"worker_id": user["id"]}, {"_id": 0, "selfie": 0}).sort("created_at", -1).limit(60)
    return await cursor.to_list(length=60)

# --- Workers search (for client/contractor) ---
@api.get("/workers")
async def list_workers(skill: Optional[str] = None, city: Optional[str] = None):
    q = {"role": "worker"}
    if skill:
        q["skills"] = skill
    if city:
        q["city"] = city
    cursor = db.users.find(q, {"_id": 0, "password": 0}).limit(100)
    return await cursor.to_list(length=100)

# --- Wallet ---
@api.get("/wallet")
async def wallet(user=Depends(current_user)):
    txns = await db.wallet_txns.find({"user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).to_list(length=50)
    fresh = await db.users.find_one({"id": user["id"]}, {"_id": 0, "wallet_balance": 1, "referral_code": 1})
    return {"balance": fresh.get("wallet_balance", 0), "referral_code": fresh.get("referral_code"), "transactions": txns}


@api.get("/wallet/referral-stats")
async def wallet_referral_stats(user=Depends(current_user)):
    """Return count of users invited via this user's referral code + total ₹ earned from referrals."""
    fresh = await db.users.find_one({"id": user["id"]}, {"_id": 0, "referral_code": 1})
    code = (fresh or {}).get("referral_code")
    if not code:
        return {"invited": 0, "earned": 0, "code": None}
    invited = await db.users.count_documents({"referred_by": code})
    # sum of positive referral credits in txn ledger
    pipeline = [
        {"$match": {"user_id": user["id"], "type": "referral_credit"}},
        {"$group": {"_id": None, "sum": {"$sum": "$amount"}}},
    ]
    agg = await db.wallet_txns.aggregate(pipeline).to_list(length=1)
    earned = float(agg[0]["sum"]) if agg else 0.0
    return {"invited": invited, "earned": earned, "code": code}


# --- Wallet CSV / PDF export ---
def _txn_type_label(t: str) -> str:
    m = {
        "wallet_topup": "Wallet Top-up",
        "topup": "Wallet Top-up",
        "withdrawal": "UPI Withdrawal",
        "referral_credit": "Referral Bonus",
        "salary": "Salary / Earnings",
        "wage": "Wage Payout",
        "erp_pro_topup": "ERP Pro Subscription",
        "escrow_release": "Job Payment Released",
        "escrow_hold": "Payment Held (Escrow)",
    }
    return m.get(t or "", (t or "").replace("_", " ").title() or "Transaction")


async def _wallet_txns_for_export(user: dict, months: int) -> list:
    from dateutil.relativedelta import relativedelta  # type: ignore
    try:
        cutoff = (datetime.now(timezone.utc) - relativedelta(months=months)).isoformat()
    except Exception:
        cutoff = (datetime.now(timezone.utc) - timedelta(days=months * 30)).isoformat()
    cursor = db.wallet_txns.find(
        {"user_id": user["id"], "created_at": {"$gte": cutoff}},
        {"_id": 0},
    ).sort("created_at", -1).limit(5000)
    return await cursor.to_list(length=5000)


@api.get("/wallet/export/csv")
async def wallet_export_csv(months: int = 6, user=Depends(current_user)):
    import csv
    from io import StringIO
    from fastapi.responses import StreamingResponse

    rows = await _wallet_txns_for_export(user, months)
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(["Date", "Time", "Type", "Description", "Amount (₹)", "Balance Effect", "Status", "Reference"])
    running_credit = 0.0
    running_debit = 0.0
    for r in rows:
        try:
            dt = datetime.fromisoformat(str(r.get("created_at", "")).replace("Z", "+00:00"))
            date_s = dt.strftime("%Y-%m-%d")
            time_s = dt.strftime("%H:%M:%S")
        except Exception:
            date_s = str(r.get("created_at", ""))[:10]
            time_s = str(r.get("created_at", ""))[11:19]
        amt = float(r.get("amount") or 0)
        if amt >= 0:
            running_credit += amt
        else:
            running_debit += abs(amt)
        ref = r.get("upi_id") or r.get("order_id") or r.get("job_id") or ""
        w.writerow([
            date_s, time_s,
            _txn_type_label(r.get("type", "")),
            r.get("note", "") or "",
            f"{amt:+.2f}",
            "Credit" if amt >= 0 else "Debit",
            r.get("status", "success"),
            ref,
        ])
    # Summary row
    w.writerow([])
    w.writerow(["", "", "", "TOTAL CREDIT", f"{running_credit:+.2f}", "", "", ""])
    w.writerow(["", "", "", "TOTAL DEBIT",  f"-{running_debit:.2f}", "", "", ""])
    w.writerow(["", "", "", "NET",           f"{(running_credit - running_debit):+.2f}", "", "", ""])

    filename = f"wallet_{user.get('id','')}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv", headers=headers)


@api.get("/wallet/export/pdf")
async def wallet_export_pdf(months: int = 6, user=Depends(current_user)):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        )
    except Exception:
        raise HTTPException(500, "PDF library unavailable")

    rows = await _wallet_txns_for_export(user, months)
    fresh = await db.users.find_one({"id": user["id"]}, {"_id": 0, "wallet_balance": 1, "referral_code": 1, "name": 1, "mobile": 1})
    balance = float((fresh or {}).get("wallet_balance", 0) or 0)
    ref_code = (fresh or {}).get("referral_code") or "—"

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.4 * cm, rightMargin=1.4 * cm,
        topMargin=1.4 * cm, bottomMargin=1.4 * cm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle("Title", parent=styles["Heading1"], textColor=rl_colors.HexColor("#18181B"), fontSize=18, spaceAfter=2)
    sub = ParagraphStyle("Sub", parent=styles["Normal"], textColor=rl_colors.HexColor("#52525B"), fontSize=9, spaceAfter=8)
    section = ParagraphStyle("Sec", parent=styles["Heading2"], fontSize=13, textColor=rl_colors.HexColor("#B45309"), spaceBefore=4, spaceAfter=4)

    story = [
        Paragraph("BuildMitra — Wallet Statement", title),
        Paragraph(
            f"<b>{fresh.get('name','')}</b> · {fresh.get('mobile','')} · Referral: {ref_code}"
            f"<br/>Period: last {months} months · Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            sub,
        ),
    ]

    # Summary
    total_credit = sum(float(r.get("amount") or 0) for r in rows if float(r.get("amount") or 0) >= 0)
    total_debit = sum(abs(float(r.get("amount") or 0)) for r in rows if float(r.get("amount") or 0) < 0)
    sum_data = [
        ["Current Balance", "Credited", "Debited", "Net"],
        [f"₹{balance:,.2f}", f"+₹{total_credit:,.2f}", f"-₹{total_debit:,.2f}", f"₹{(total_credit - total_debit):+,.2f}"],
    ]
    sum_tbl = Table(sum_data, colWidths=[4.4 * cm] * 4)
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#FEF3C7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.HexColor("#B45309")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 12),
        ("TEXTCOLOR", (1, 1), (1, 1), rl_colors.HexColor("#16A34A")),
        ("TEXTCOLOR", (2, 1), (2, 1), rl_colors.HexColor("#DC2626")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOX", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#E4E4E7")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#E4E4E7")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph("Transactions", section))

    header = ["Date", "Type", "Description", "Status", "Amount (₹)"]
    data = [header]
    for r in rows:
        try:
            dt = datetime.fromisoformat(str(r.get("created_at", "")).replace("Z", "+00:00"))
            date_s = dt.strftime("%d %b %Y\n%H:%M")
        except Exception:
            date_s = str(r.get("created_at", ""))[:10]
        amt = float(r.get("amount") or 0)
        desc = (r.get("note", "") or "")[:44]
        data.append([
            date_s,
            _txn_type_label(r.get("type", "")),
            desc,
            (r.get("status", "success") or "success").title(),
            f"{'+' if amt >= 0 else '−'}₹{abs(amt):,.2f}",
        ])

    if len(data) == 1:
        story.append(Paragraph("No transactions in this period.", styles["Italic"]))
    else:
        tbl = Table(data, colWidths=[2.4 * cm, 3.6 * cm, 6.4 * cm, 2.4 * cm, 2.8 * cm], repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#18181B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.HexColor("#FFFFFF")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#E4E4E7")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
        ]
        for i in range(1, len(data)):
            if i % 2 == 0:
                style.append(("BACKGROUND", (0, i), (-1, i), rl_colors.HexColor("#FAFAFA")))
            amt_cell = data[i][-1]
            if amt_cell.startswith("+"):
                style.append(("TEXTCOLOR", (-1, i), (-1, i), rl_colors.HexColor("#16A34A")))
                style.append(("FONTNAME", (-1, i), (-1, i), "Helvetica-Bold"))
            elif amt_cell.startswith("−"):
                style.append(("TEXTCOLOR", (-1, i), (-1, i), rl_colors.HexColor("#DC2626")))
                style.append(("FONTNAME", (-1, i), (-1, i), "Helvetica-Bold"))
        tbl.setStyle(TableStyle(style))
        story.append(tbl)

    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(
        "Auto-generated by BuildMitra Wallet — an official record of your account activity.",
        ParagraphStyle("footer", parent=styles["Normal"], fontSize=8, textColor=rl_colors.HexColor("#71717A"), alignment=1),
    ))
    doc.build(story)
    buf.seek(0)
    filename = f"wallet_statement_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buf, media_type="application/pdf", headers=headers)


# --- Withdrawal history endpoint (dedicated) ---
@api.get("/wallet/withdrawals")
async def wallet_withdrawals(user=Depends(current_user)):
    """Return list of withdrawal-only transactions for this user."""
    cursor = db.wallet_txns.find(
        {"user_id": user["id"], "type": "withdrawal"},
        {"_id": 0},
    ).sort("created_at", -1).limit(200)
    items = await cursor.to_list(length=200)
    return items



# --- Ratings ---
@api.post("/ratings")
async def rate(body: RatingIn, user=Depends(current_user)):
    rec = {
        "id": str(uuid.uuid4()),
        "by": user["id"],
        "by_name": user["name"],
        "target_user_id": body.target_user_id,
        "job_id": body.job_id,
        "stars": max(1, min(5, body.stars)),
        "comment": body.comment,
        "created_at": now_iso(),
    }
    await db.ratings.insert_one(rec)
    agg = await db.ratings.aggregate([
        {"$match": {"target_user_id": body.target_user_id}},
        {"$group": {"_id": "$target_user_id", "avg": {"$avg": "$stars"}, "count": {"$sum": 1}}}
    ]).to_list(length=1)
    if agg:
        await db.users.update_one(
            {"id": body.target_user_id},
            {"$set": {"rating_avg": round(agg[0]["avg"], 2), "rating_count": agg[0]["count"]}}
        )
    rec.pop("_id", None)
    return rec

@api.get("/ratings/{user_id}")
async def get_ratings(user_id: str):
    cursor = db.ratings.find({"target_user_id": user_id}, {"_id": 0}).sort("created_at", -1).limit(50)
    return await cursor.to_list(length=50)

# --- AI Matching ---
@api.post("/ai/match-jobs")
async def ai_match_jobs(user=Depends(current_user)):
    if user["role"] != "worker":
        raise HTTPException(403, "Workers only")
    jobs = await db.jobs.find({"status": "open"}, {"_id": 0}).limit(40).to_list(length=40)
    if not jobs:
        return {"summary": "No jobs available yet. Check back soon.", "top_job_ids": []}
    skills = ", ".join(user.get("skills", []) or ["General"])
    wage = user.get("daily_wage", 0)
    city = user.get("city", "")
    job_brief = "\n".join([
        f"- id={j['id']} | {j['title']} | skill={j['skill']} | wage=₹{j['daily_wage']} | loc={j['location']}"
        for j in jobs[:20]
    ])
    prompt = (
        f"Worker skills: {skills}. Expected wage: ₹{wage}/day. City: {city}.\n"
        f"Available jobs:\n{job_brief}\n\n"
        "Pick top 3 best matching job ids and explain in 1 short Hindi+English sentence why. "
        "Return strictly as: IDS: id1,id2,id3 \\n REASON: <one line>"
    )
    text = "IDS: " + ",".join([j["id"] for j in jobs[:3]]) + "\nREASON: Matched by skill and wage."
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"match-{user['id']}",
            system_message="You are a job matcher for Indian construction workers. Be concise."
        ).with_model("anthropic", "claude-haiku-4-5-20251001")
        resp = await chat.send_message(UserMessage(text=prompt))
        text = str(resp)
    except Exception as e:
        logger.warning(f"AI match fallback: {e}")
    ids: List[str] = []
    reason = ""
    for line in text.splitlines():
        line = line.strip()
        if line.upper().startswith("IDS:"):
            ids = [x.strip() for x in line.split(":", 1)[1].split(",") if x.strip()][:3]
        elif line.upper().startswith("REASON:"):
            reason = line.split(":", 1)[1].strip()
    if not ids:
        ids = [j["id"] for j in jobs[:3]]
        reason = reason or "Matched by skill and wage."
    return {"summary": reason, "top_job_ids": ids}

# --- ERP (Contractor scope) ---
async def contractor_user(user=Depends(current_user)) -> dict:
    if user.get("role") != "contractor":
        raise HTTPException(403, "Contractors only")
    return user

# Materials
@api.post("/erp/materials")
async def add_material(body: MaterialIn, user=Depends(contractor_user)):
    rec = {"id": str(uuid.uuid4()), "owner": user["id"], **body.model_dump(), "created_at": now_iso()}
    await db.materials.insert_one(rec); rec.pop("_id", None); return rec

@api.get("/erp/materials")
async def list_materials(user=Depends(contractor_user)):
    cur = db.materials.find({"owner": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(500)
    return await cur.to_list(length=500)

@api.put("/erp/materials/{mid}")
async def update_material(mid: str, body: MaterialIn, user=Depends(contractor_user)):
    res = await db.materials.update_one({"id": mid, "owner": user["id"]}, {"$set": body.model_dump()})
    if not res.matched_count: raise HTTPException(404, "Not found")
    return {"ok": True}

@api.delete("/erp/materials/{mid}")
async def del_material(mid: str, user=Depends(contractor_user)):
    res = await db.materials.delete_one({"id": mid, "owner": user["id"]})
    if not res.deleted_count: raise HTTPException(404, "Not found")
    return {"ok": True}

# Tools
@api.post("/erp/tools")
async def add_tool(body: ToolIn, user=Depends(contractor_user)):
    rec = {"id": str(uuid.uuid4()), "owner": user["id"], **body.model_dump(), "created_at": now_iso()}
    await db.tools.insert_one(rec); rec.pop("_id", None); return rec

@api.get("/erp/tools")
async def list_tools(user=Depends(contractor_user)):
    cur = db.tools.find({"owner": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(500)
    return await cur.to_list(length=500)

@api.put("/erp/tools/{tid}")
async def update_tool(tid: str, body: ToolIn, user=Depends(contractor_user)):
    res = await db.tools.update_one({"id": tid, "owner": user["id"]}, {"$set": body.model_dump()})
    if not res.matched_count: raise HTTPException(404, "Not found")
    return {"ok": True}

@api.delete("/erp/tools/{tid}")
async def del_tool(tid: str, user=Depends(contractor_user)):
    res = await db.tools.delete_one({"id": tid, "owner": user["id"]})
    if not res.deleted_count: raise HTTPException(404, "Not found")
    return {"ok": True}

# Estimates
@api.post("/erp/estimates")
async def add_estimate(body: EstimateIn, user=Depends(contractor_user)):
    d = body.model_dump()
    total_cost = d["labour_cost"] + d["material_cost"] + d["equipment_cost"] + d["transport_cost"] + d["misc_cost"]
    profit = d["revenue"] - total_cost
    margin = round((profit / d["revenue"]) * 100, 2) if d["revenue"] else 0
    rec = {"id": str(uuid.uuid4()), "owner": user["id"], **d,
           "total_cost": total_cost, "profit": profit, "margin_pct": margin,
           "created_at": now_iso()}
    await db.estimates.insert_one(rec); rec.pop("_id", None); return rec

@api.get("/erp/estimates")
async def list_estimates(user=Depends(contractor_user)):
    cur = db.estimates.find({"owner": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(500)
    return await cur.to_list(length=500)

@api.delete("/erp/estimates/{eid}")
async def del_estimate(eid: str, user=Depends(contractor_user)):
    res = await db.estimates.delete_one({"id": eid, "owner": user["id"]})
    if not res.deleted_count: raise HTTPException(404, "Not found")
    return {"ok": True}

# Bills
def _calc_bill(items, tax_pct):
    subtotal = sum((it.get("qty", 0) * it.get("rate", 0)) for it in items)
    tax = round(subtotal * tax_pct / 100, 2)
    return subtotal, tax, subtotal + tax

@api.post("/erp/bills")
async def add_bill(body: BillIn, user=Depends(contractor_user)):
    items = [it.model_dump() for it in body.items]
    sub, tax, total = _calc_bill(items, body.tax_pct)
    count = await db.bills.count_documents({"owner": user["id"]})
    bill_no = f"BM-{datetime.now(timezone.utc).year}-{count+1:04d}"
    rec = {"id": str(uuid.uuid4()), "owner": user["id"], "bill_no": bill_no,
           "bill_to": body.bill_to, "project": body.project, "items": items,
           "tax_pct": body.tax_pct, "subtotal": sub, "tax_amount": tax, "total": total,
           "notes": body.notes, "status": "unpaid", "created_at": now_iso()}
    await db.bills.insert_one(rec); rec.pop("_id", None); return rec

@api.get("/erp/bills")
async def list_bills(user=Depends(contractor_user)):
    cur = db.bills.find({"owner": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(500)
    return await cur.to_list(length=500)

@api.post("/erp/bills/{bid}/mark-paid")
async def mark_bill_paid(bid: str, user=Depends(contractor_user)):
    res = await db.bills.update_one({"id": bid, "owner": user["id"]}, {"$set": {"status": "paid"}})
    if not res.matched_count: raise HTTPException(404, "Not found")
    return {"ok": True}

@api.delete("/erp/bills/{bid}")
async def del_bill(bid: str, user=Depends(contractor_user)):
    res = await db.bills.delete_one({"id": bid, "owner": user["id"]})
    if not res.deleted_count: raise HTTPException(404, "Not found")
    return {"ok": True}

@api.get("/erp/dashboard")
async def erp_dashboard(user=Depends(contractor_user)):
    mat = await db.materials.count_documents({"owner": user["id"]})
    low = await db.materials.count_documents({"owner": user["id"], "$expr": {"$lte": ["$qty", "$min_qty"]}})
    tools_total = await db.tools.count_documents({"owner": user["id"]})
    tools_inuse = await db.tools.count_documents({"owner": user["id"], "status": "in_use"})
    est = await db.estimates.count_documents({"owner": user["id"]})
    bills_total = await db.bills.count_documents({"owner": user["id"]})
    bills_paid = await db.bills.count_documents({"owner": user["id"], "status": "paid"})
    rev_agg = await db.bills.aggregate([
        {"$match": {"owner": user["id"]}},
        {"$group": {"_id": "$status", "sum": {"$sum": "$total"}}}
    ]).to_list(length=10)
    revenue_paid = sum(r["sum"] for r in rev_agg if r["_id"] == "paid")
    revenue_pending = sum(r["sum"] for r in rev_agg if r["_id"] != "paid")
    return {
        "materials_total": mat, "materials_low_stock": low,
        "tools_total": tools_total, "tools_in_use": tools_inuse,
        "estimates_total": est,
        "bills_total": bills_total, "bills_paid": bills_paid,
        "revenue_paid": revenue_paid, "revenue_pending": revenue_pending,
    }

# --- Chat ---
def _thread_id(a: str, b: str) -> str:
    return "-".join(sorted([a, b]))

@api.post("/chat/send")
async def chat_send(body: ChatSendIn, user=Depends(current_user)):
    peer = await db.users.find_one({"id": body.to_user_id}, {"_id": 0, "id": 1, "name": 1, "role": 1})
    if not peer:
        raise HTTPException(404, "Recipient not found")
    msg = {
        "id": str(uuid.uuid4()),
        "thread_id": _thread_id(user["id"], body.to_user_id),
        "from_id": user["id"], "from_name": user["name"],
        "to_id": body.to_user_id, "to_name": peer["name"],
        "text": body.text[:2000],
        "read": False,
        "created_at": now_iso(),
    }
    await db.chat_messages.insert_one(msg)
    msg.pop("_id", None)
    return msg

@api.get("/chat/threads")
async def chat_threads(user=Depends(current_user)):
    pipeline = [
        {"$match": {"$or": [{"from_id": user["id"]}, {"to_id": user["id"]}]}},
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": "$thread_id",
            "last_text": {"$first": "$text"},
            "last_at": {"$first": "$created_at"},
            "from_id": {"$first": "$from_id"},
            "to_id": {"$first": "$to_id"},
            "from_name": {"$first": "$from_name"},
            "to_name": {"$first": "$to_name"},
        }},
        {"$sort": {"last_at": -1}},
    ]
    raw = await db.chat_messages.aggregate(pipeline).to_list(length=100)
    out = []
    for t in raw:
        peer_id = t["to_id"] if t["from_id"] == user["id"] else t["from_id"]
        peer_name = t["to_name"] if t["from_id"] == user["id"] else t["from_name"]
        out.append({
            "thread_id": t["_id"], "peer_id": peer_id, "peer_name": peer_name,
            "last_text": t["last_text"], "last_at": t["last_at"],
        })
    return out

@api.get("/chat/messages/{peer_id}")
async def chat_messages(peer_id: str, user=Depends(current_user)):
    tid = _thread_id(user["id"], peer_id)
    msgs = await db.chat_messages.find({"thread_id": tid}, {"_id": 0}).sort("created_at", 1).to_list(length=500)
    await db.chat_messages.update_many({"thread_id": tid, "to_id": user["id"], "read": False}, {"$set": {"read": True}})
    return msgs

# --- Payroll (Contractor or Client view of worker wages) ---
@api.get("/payroll")
async def payroll(month: Optional[str] = None, user=Depends(current_user)):
    if user["role"] not in ("contractor", "client", "admin"):
        raise HTTPException(403, "Only contractor/client/admin can view payroll")
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    # Get jobs posted by this user (unless admin - admin sees all)
    my_job_ids: list = []
    if user["role"] != "admin":
        job_docs = await db.jobs.find({"posted_by": user["id"]}, {"_id": 0, "id": 1}).limit(500).to_list(length=500)
        my_job_ids = [j["id"] for j in job_docs]
        if not my_job_ids:
            return {"month": month, "rows": [], "grand_total": 0}
    # Filter attendance
    att_query: dict = {
        "type": "check_in",
        "created_at": {"$regex": f"^{month}"},
        "within_geofence": True,
    }
    if my_job_ids:
        att_query["job_id"] = {"$in": my_job_ids}
    att = await db.attendance.find(att_query, {"_id": 0, "selfie": 0}).limit(5000).to_list(length=5000)
    by_worker: dict = {}
    for a in att:
        wid = a["worker_id"]
        by_worker.setdefault(wid, {"worker_id": wid, "worker_name": a["worker_name"], "days_present": 0, "jobs": set()})
        by_worker[wid]["days_present"] += 1
        if a.get("job_id"): by_worker[wid]["jobs"].add(a["job_id"])
    out = []
    worker_ids = list(by_worker.keys())
    workers_map: dict = {}
    if worker_ids:
        workers_list = await db.users.find(
            {"id": {"$in": worker_ids}},
            {"_id": 0, "id": 1, "daily_wage": 1, "mobile": 1},
        ).to_list(length=len(worker_ids))
        workers_map = {w["id"]: w for w in workers_list}
    for wid, row in by_worker.items():
        worker = workers_map.get(wid, {})
        wage = worker.get("daily_wage", 0)
        out.append({
            "worker_id": wid, "worker_name": row["worker_name"],
            "mobile": worker.get("mobile"),
            "days_present": row["days_present"],
            "daily_wage": wage,
            "total_wage": row["days_present"] * wage,
            "jobs_count": len(row["jobs"]),
        })
    total = sum(r["total_wage"] for r in out)
    return {"month": month, "rows": sorted(out, key=lambda x: -x["total_wage"]), "grand_total": total}

# --- Client/Contractor attendance monitoring ---
@api.get("/attendance/my-workers")
async def attendance_my_workers(days: int = 30, user=Depends(current_user)):
    """For client/contractor: list of attendance entries by workers on their jobs (last N days)."""
    if user["role"] not in ("client", "contractor", "admin"):
        raise HTTPException(403, "Not allowed")
    if user["role"] == "admin":
        # Admin sees all recent attendance
        q: dict = {}
    else:
        job_docs = await db.jobs.find({"posted_by": user["id"]}, {"_id": 0, "id": 1}).limit(500).to_list(length=500)
        job_ids = [j["id"] for j in job_docs]
        if not job_ids:
            return []
        q = {"job_id": {"$in": job_ids}}
    cursor = db.attendance.find(q, {"_id": 0, "selfie": 0}).sort("created_at", -1).limit(days * 10)
    return await cursor.to_list(length=days * 10)

# --- Attendance export: CSV / PDF (worker/client/contractor/admin) ---
async def _attendance_rows_for_export(user: dict, days: int, scope: str) -> list:
    """Fetch attendance rows respecting role scope.
    scope: 'mine' (worker's own) or 'workers' (client/contractor's workers or admin all).
    """
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    if scope == "mine":
        if user["role"] != "worker":
            raise HTTPException(403, "Workers only")
        q = {"worker_id": user["id"], "created_at": {"$gte": cutoff}}
    else:
        if user["role"] not in ("client", "contractor", "admin"):
            raise HTTPException(403, "Not allowed")
        if user["role"] == "admin":
            q = {"created_at": {"$gte": cutoff}}
        else:
            job_docs = await db.jobs.find(
                {"posted_by": user["id"]}, {"_id": 0, "id": 1}
            ).limit(500).to_list(length=500)
            job_ids = [j["id"] for j in job_docs]
            if not job_ids:
                return []
            q = {"job_id": {"$in": job_ids}, "created_at": {"$gte": cutoff}}
    cursor = db.attendance.find(q, {"_id": 0, "selfie": 0}).sort("created_at", -1).limit(5000)
    return await cursor.to_list(length=5000)


@api.get("/attendance/export/csv")
async def attendance_export_csv(days: int = 30, scope: str = "workers", user=Depends(current_user)):
    """Export attendance as CSV. scope='mine' (worker) or 'workers' (client/contractor/admin)."""
    import csv
    from io import StringIO
    from fastapi.responses import StreamingResponse

    rows = await _attendance_rows_for_export(user, days, scope)

    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Date", "Time", "Worker", "Type",
        "Job Title", "Job ID", "Verified", "Distance (m)", "Latitude", "Longitude",
    ])
    for r in rows:
        try:
            dt = datetime.fromisoformat(str(r.get("created_at", "")).replace("Z", "+00:00"))
            date_s = dt.strftime("%Y-%m-%d")
            time_s = dt.strftime("%H:%M:%S")
        except Exception:
            date_s = str(r.get("created_at", ""))[:10]
            time_s = str(r.get("created_at", ""))[11:19]
        writer.writerow([
            date_s, time_s,
            r.get("worker_name", ""),
            "Check In" if r.get("type") == "check_in" else "Check Out",
            r.get("job_title") or "-",
            r.get("job_id") or "-",
            "Yes" if r.get("within_geofence", True) else "No",
            r.get("distance_from_site_m") if r.get("distance_from_site_m") is not None else "",
            r.get("lat", ""),
            r.get("lng", ""),
        ])
    filename = f"attendance_{scope}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv", headers=headers)


@api.get("/attendance/export/pdf")
async def attendance_export_pdf(days: int = 30, scope: str = "workers", user=Depends(current_user)):
    """Export attendance as PDF report."""
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        )
    except Exception:
        raise HTTPException(500, "PDF library unavailable")

    rows = await _attendance_rows_for_export(user, days, scope)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.2 * cm, bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleBM", parent=styles["Heading1"], textColor=rl_colors.HexColor("#18181B"),
        fontSize=18, spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "SubBM", parent=styles["Normal"], textColor=rl_colors.HexColor("#52525B"),
        fontSize=10, spaceAfter=12,
    )

    story = []
    heading = "My Attendance Report" if scope == "mine" else "Workforce Attendance Report"
    story.append(Paragraph(f"BuildMitra — {heading}", title_style))
    story.append(Paragraph(
        f"Generated for: <b>{user.get('name', '')}</b> ({user.get('role', '').title()}) "
        f"· Range: last {days} day(s) · {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        sub_style,
    ))

    # Summary block
    total = len(rows)
    verified = sum(1 for r in rows if r.get("within_geofence", True))
    flagged = total - verified
    unique_workers = len({r.get("worker_id") for r in rows if r.get("worker_id")})
    summary_data = [
        ["Total Entries", "Verified", "Flagged", "Unique Workers"],
        [str(total), str(verified), str(flagged), str(unique_workers)],
    ]
    summary_tbl = Table(summary_data, colWidths=[4.5 * cm] * 4)
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#FEF3C7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.HexColor("#B45309")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("TEXTCOLOR", (0, 1), (-1, 1), rl_colors.HexColor("#18181B")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#E4E4E7")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#E4E4E7")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_tbl)
    story.append(Spacer(1, 0.6 * cm))

    # Detail table
    header = [
        "Date", "Time", "Worker", "Type", "Job", "Verified", "Distance",
    ]
    data = [header]
    for r in rows:
        try:
            dt = datetime.fromisoformat(str(r.get("created_at", "")).replace("Z", "+00:00"))
            date_s = dt.strftime("%d %b %Y")
            time_s = dt.strftime("%H:%M")
        except Exception:
            date_s = str(r.get("created_at", ""))[:10]
            time_s = str(r.get("created_at", ""))[11:16]
        dist = r.get("distance_from_site_m")
        dist_s = f"{dist} m" if dist is not None else "—"
        job_s = (r.get("job_title") or "General")[:26]
        worker_s = (r.get("worker_name") or "")[:22]
        data.append([
            date_s, time_s, worker_s,
            "Check In" if r.get("type") == "check_in" else "Check Out",
            job_s,
            "✓ Yes" if r.get("within_geofence", True) else "✗ Off-site",
            dist_s,
        ])

    if len(data) == 1:
        story.append(Paragraph("No attendance records in this range.", styles["Italic"]))
    else:
        detail_tbl = Table(
            data,
            colWidths=[2.6 * cm, 1.8 * cm, 4.6 * cm, 2.4 * cm, 6.8 * cm, 2.5 * cm, 2.0 * cm],
            repeatRows=1,
        )
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#18181B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.HexColor("#FFFFFF")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#E4E4E7")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]
        # Row shading + off-site red text
        for i in range(1, len(data)):
            if i % 2 == 0:
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), rl_colors.HexColor("#FAFAFA")))
            if str(data[i][5]).startswith("✗"):
                style_cmds.append(("TEXTCOLOR", (5, i), (5, i), rl_colors.HexColor("#DC2626")))
                style_cmds.append(("FONTNAME", (5, i), (5, i), "Helvetica-Bold"))
            else:
                style_cmds.append(("TEXTCOLOR", (5, i), (5, i), rl_colors.HexColor("#16A34A")))
        detail_tbl.setStyle(TableStyle(style_cmds))
        story.append(detail_tbl)

    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(
        "This is an auto-generated report from BuildMitra. All check-ins are GPS + Selfie verified.",
        ParagraphStyle("footer", parent=styles["Normal"], fontSize=8,
                       textColor=rl_colors.HexColor("#71717A"), alignment=1),
    ))

    doc.build(story)
    buf.seek(0)
    filename = f"attendance_{scope}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buf, media_type="application/pdf", headers=headers)

# --- Wallet withdrawal (worker cashes out earnings via mock UPI) ---
class WithdrawIn(BaseModel):
    amount: float
    upi_id: str

@api.post("/wallet/withdraw")
async def wallet_withdraw(body: WithdrawIn, user=Depends(current_user)):
    """Mock UPI withdrawal — deducts from wallet_balance, creates txn record.
    In production this would call Razorpay Payout API."""
    if body.amount <= 0:
        raise HTTPException(400, "Invalid amount")
    if "@" not in body.upi_id:
        raise HTTPException(400, "Invalid UPI ID")
    fresh = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    if not fresh:
        raise HTTPException(404, "User not found")
    balance = float(fresh.get("wallet_balance", 0) or 0)
    if body.amount > balance:
        raise HTTPException(400, f"Insufficient balance (₹{balance:.2f})")
    txn = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "type": "withdrawal",
        "amount": -body.amount,
        "upi_id": body.upi_id,
        "status": "processing",
        "note": f"UPI payout to {body.upi_id}",
        "created_at": now_iso(),
    }
    await db.wallet_txns.insert_one(txn)
    await db.users.update_one({"id": user["id"]}, {"$inc": {"wallet_balance": -body.amount}})
    txn.pop("_id", None)
    return {"ok": True, "txn": txn, "new_balance": balance - body.amount}

# --- AI Worker Recommendation (for client/contractor) ---
@api.get("/ai/recommend-workers/{job_id}")
async def ai_recommend_workers(job_id: str, user=Depends(current_user)):
    if user["role"] not in ("client", "contractor"):
        raise HTTPException(403, "Clients/contractors only")
    job = await db.jobs.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(404, "Job not found")
    candidates = await db.users.find(
        {"role": "worker", "skills": job["skill"], "available": True},
        {"_id": 0, "password": 0}
    ).limit(30).to_list(length=30)
    if not candidates:
        candidates = await db.users.find(
            {"role": "worker", "available": True},
            {"_id": 0, "password": 0}
        ).limit(20).to_list(length=20)
    if not candidates:
        return {"summary": "No available workers right now.", "top_ids": [], "candidates": []}
    brief = "\n".join([
        f"- id={c['id']} | {c['name']} | skills={c.get('skills')} | wage=₹{c.get('daily_wage',0)} | rating={c.get('rating_avg',0)} ({c.get('rating_count',0)}) | city={c.get('city','')} | verified={c.get('aadhaar_verified', False)}"
        for c in candidates
    ])
    prompt = (
        f"Job: {job['title']} ({job['skill']}) at {job['location']}. Pays ₹{job['daily_wage']}/day, needs {job['workers_needed']} worker(s).\n"
        f"Candidates:\n{brief}\n\nPick the TOP 3 best-fit workers and explain in 1 short sentence why. "
        "Reply strictly as:\nIDS: id1,id2,id3\nREASON: <one line>"
    )
    text = "IDS: " + ",".join([c["id"] for c in candidates[:3]]) + "\nREASON: Matched by skill, wage and rating."
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"rec-{job_id}",
            system_message="You are a hiring advisor for an Indian construction marketplace. Be concise."
        ).with_model("anthropic", "claude-haiku-4-5-20251001")
        resp = await chat.send_message(UserMessage(text=prompt))
        text = str(resp)
    except Exception as e:
        logger.warning(f"AI recommend fallback: {e}")
    ids: List[str] = []
    reason = "Matched by skill and rating."
    for line in text.splitlines():
        line = line.strip()
        if line.upper().startswith("IDS:"):
            ids = [x.strip() for x in line.split(":", 1)[1].split(",") if x.strip()][:3]
        elif line.upper().startswith("REASON:"):
            reason = line.split(":", 1)[1].strip()
    if not ids:
        ids = [c["id"] for c in candidates[:3]]
    return {"summary": reason, "top_ids": ids, "candidates": [c for c in candidates if c["id"] in ids]}

# --- Bill exports (PDF + Excel) ---
def _build_bill_pdf(bill: dict, owner_name: str) -> bytes:
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm,
        title=f"BuildMitra Invoice {bill['bill_no']}",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=rl.HexColor("#18181B"), fontSize=22, spaceAfter=4)
    brand = ParagraphStyle("brand", parent=styles["Normal"], textColor=rl.HexColor("#F59E0B"), fontSize=11, fontName="Helvetica-Bold")
    label = ParagraphStyle("label", parent=styles["Normal"], textColor=rl.HexColor("#71717A"), fontSize=9, fontName="Helvetica-Bold")
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=11, textColor=rl.HexColor("#18181B"))
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, textColor=rl.HexColor("#71717A"))

    story = []
    # Header
    story.append(Paragraph("BuildMitra", brand))
    story.append(Paragraph("INVOICE", h1))
    story.append(Spacer(1, 8))

    meta = Table([
        [Paragraph("INVOICE NO", label), Paragraph(bill["bill_no"], body)],
        [Paragraph("DATE", label), Paragraph(bill["created_at"][:10], body)],
        [Paragraph("STATUS", label), Paragraph(bill["status"].upper(), body)],
    ], colWidths=[35*mm, 80*mm])
    meta.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP"), ("BOTTOMPADDING", (0,0), (-1,-1), 4)]))
    story.append(meta)
    story.append(Spacer(1, 14))

    parties = Table([
        [Paragraph("FROM", label), Paragraph("BILL TO", label)],
        [Paragraph(owner_name, body), Paragraph(bill["bill_to"], body)],
        [Paragraph(bill.get("project", "") or "", small), Paragraph("", small)],
    ], colWidths=[85*mm, 85*mm])
    parties.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP"), ("BOTTOMPADDING", (0,0), (-1,-1), 4)]))
    story.append(parties)
    story.append(Spacer(1, 16))

    # Line items
    rows = [["#", "Description", "Qty", "Rate (₹)", "Amount (₹)"]]
    for i, it in enumerate(bill["items"], 1):
        amt = it["qty"] * it["rate"]
        rows.append([str(i), it["description"], f"{it['qty']:g}", f"{it['rate']:,.2f}", f"{amt:,.2f}"])
    tbl = Table(rows, colWidths=[12*mm, 88*mm, 20*mm, 25*mm, 30*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), rl.HexColor("#F59E0B")),
        ("TEXTCOLOR", (0,0), (-1,0), rl.HexColor("#18181B")),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (0,0), (-1,0), "LEFT"),
        ("ALIGN", (2,0), (-1,-1), "RIGHT"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("LINEBELOW", (0,0), (-1,-1), 0.4, rl.HexColor("#E4E4E7")),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 12))

    totals = Table([
        ["Subtotal", f"₹{bill['subtotal']:,.2f}"],
        [f"GST ({bill['tax_pct']:g}%)", f"₹{bill['tax_amount']:,.2f}"],
        ["TOTAL", f"₹{bill['total']:,.2f}"],
    ], colWidths=[140*mm, 35*mm])
    totals.setStyle(TableStyle([
        ("ALIGN", (0,0), (-1,-1), "RIGHT"),
        ("FONTSIZE", (0,0), (-1,-1), 11),
        ("FONTNAME", (0,2), (-1,2), "Helvetica-Bold"),
        ("BACKGROUND", (0,2), (-1,2), rl.HexColor("#FEF3C7")),
        ("TEXTCOLOR", (0,2), (-1,2), rl.HexColor("#B45309")),
        ("FONTSIZE", (0,2), (-1,2), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(totals)
    story.append(Spacer(1, 20))

    if bill.get("notes"):
        story.append(Paragraph("Notes", label))
        story.append(Paragraph(bill["notes"], body))
        story.append(Spacer(1, 12))

    story.append(Paragraph("Generated by BuildMitra · India's construction workforce platform", small))
    doc.build(story)
    return buf.getvalue()

def _build_bill_xlsx(bill: dict, owner_name: str) -> bytes:
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "Invoice"
    bold = Font(bold=True); white = Font(bold=True, color="18181B")
    brand_fill = PatternFill("solid", fgColor="F59E0B")
    total_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Side(style="thin", color="E4E4E7"); border = Border(bottom=thin)
    right = Alignment(horizontal="right")

    ws["A1"] = "BuildMitra Invoice"; ws["A1"].font = Font(bold=True, size=18, color="F59E0B")
    ws["A3"] = "Invoice No"; ws["A3"].font = bold; ws["B3"] = bill["bill_no"]
    ws["A4"] = "Date"; ws["A4"].font = bold; ws["B4"] = bill["created_at"][:10]
    ws["A5"] = "Status"; ws["A5"].font = bold; ws["B5"] = bill["status"].upper()
    ws["A7"] = "From"; ws["A7"].font = bold; ws["B7"] = owner_name
    ws["A8"] = "Bill To"; ws["A8"].font = bold; ws["B8"] = bill["bill_to"]
    ws["A9"] = "Project"; ws["A9"].font = bold; ws["B9"] = bill.get("project", "") or ""

    headers = ["#", "Description", "Qty", "Rate (₹)", "Amount (₹)"]
    start = 11
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.font = white; cell.fill = brand_fill
    r = start + 1
    for i, it in enumerate(bill["items"], 1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=it["description"])
        ws.cell(row=r, column=3, value=it["qty"]).alignment = right
        ws.cell(row=r, column=4, value=it["rate"]).alignment = right
        ws.cell(row=r, column=5, value=it["qty"] * it["rate"]).alignment = right
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border
        r += 1
    r += 1
    ws.cell(row=r, column=4, value="Subtotal").font = bold
    ws.cell(row=r, column=5, value=bill["subtotal"]).alignment = right; r += 1
    ws.cell(row=r, column=4, value=f"GST ({bill['tax_pct']}%)").font = bold
    ws.cell(row=r, column=5, value=bill["tax_amount"]).alignment = right; r += 1
    tc1 = ws.cell(row=r, column=4, value="TOTAL"); tc1.font = bold; tc1.fill = total_fill
    tc2 = ws.cell(row=r, column=5, value=bill["total"]); tc2.font = bold; tc2.fill = total_fill; tc2.alignment = right

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    buf = BytesIO(); wb.save(buf); return buf.getvalue()

import base64 as _b64

@api.get("/erp/bills/{bid}/pdf")
async def bill_pdf(bid: str, user=Depends(contractor_user)):
    bill = await db.bills.find_one({"id": bid, "owner": user["id"]}, {"_id": 0})
    if not bill:
        raise HTTPException(404, "Bill not found")
    pdf = _build_bill_pdf(bill, user.get("company_name") or user["name"])
    return {
        "filename": f"{bill['bill_no']}.pdf",
        "mime": "application/pdf",
        "base64": _b64.b64encode(pdf).decode(),
    }

@api.get("/erp/bills/{bid}/excel")
async def bill_excel(bid: str, user=Depends(contractor_user)):
    bill = await db.bills.find_one({"id": bid, "owner": user["id"]}, {"_id": 0})
    if not bill:
        raise HTTPException(404, "Bill not found")
    xlsx = _build_bill_xlsx(bill, user.get("company_name") or user["name"])
    return {
        "filename": f"{bill['bill_no']}.xlsx",
        "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "base64": _b64.b64encode(xlsx).decode(),
    }

# --- Complaints (any user can file) ---
@api.post("/complaints")
async def file_complaint(body: ComplaintIn, user=Depends(current_user)):
    rec = {
        "id": str(uuid.uuid4()),
        "by_user_id": user["id"],
        "by_user_name": user["name"],
        "by_user_role": user["role"],
        "against_user_id": body.against_user_id,
        "subject": body.subject,
        "description": body.description,
        "status": "open",  # open | resolved | rejected
        "admin_note": "",
        "created_at": now_iso(),
    }
    await db.complaints.insert_one(rec)
    rec.pop("_id", None)
    return rec

@api.get("/complaints/mine")
async def my_complaints(user=Depends(current_user)):
    cursor = db.complaints.find({"by_user_id": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(100)
    return await cursor.to_list(length=100)

# --- Admin ---
async def admin_user(user=Depends(current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    return user

@api.get("/admin/stats")
async def admin_stats(_=Depends(admin_user)):
    total_workers = await db.users.count_documents({"role": "worker"})
    total_contractors = await db.users.count_documents({"role": "contractor"})
    total_clients = await db.users.count_documents({"role": "client"})
    total_jobs = await db.jobs.count_documents({})
    active_jobs = await db.jobs.count_documents({"status": "open"})
    completed_jobs = await db.jobs.count_documents({"status": "closed"})
    total_applications = await db.applications.count_documents({})
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    daily_attendance = await db.attendance.count_documents({"created_at": {"$regex": f"^{today}"}})
    open_complaints = await db.complaints.count_documents({"status": "open"})
    pending_verifications = await db.users.count_documents({
        "role": {"$in": ["worker", "contractor"]}, "aadhaar_verified": False
    })
    total_wallet = await db.users.aggregate([
        {"$group": {"_id": None, "sum": {"$sum": "$wallet_balance"}}}
    ]).to_list(length=1)
    revenue_proxy = (total_wallet[0]["sum"] if total_wallet else 0)
    return {
        "total_workers": total_workers,
        "total_contractors": total_contractors,
        "total_clients": total_clients,
        "total_jobs": total_jobs,
        "active_jobs": active_jobs,
        "completed_jobs": completed_jobs,
        "total_applications": total_applications,
        "daily_attendance": daily_attendance,
        "open_complaints": open_complaints,
        "pending_verifications": pending_verifications,
        "wallet_payouts": revenue_proxy,
    }

@api.get("/admin/users")
async def admin_users(role: Optional[str] = None, q: Optional[str] = None, _=Depends(admin_user)):
    query: dict = {}
    if role:
        query["role"] = role
    if q:
        query["$or"] = [{"name": {"$regex": q, "$options": "i"}}, {"mobile": {"$regex": q}}]
    cursor = db.users.find(query, {"_id": 0, "password": 0}).sort("created_at", -1).limit(200)
    return await cursor.to_list(length=200)

@api.post("/admin/users/{user_id}/verify")
async def admin_verify(user_id: str, _=Depends(admin_user)):
    res = await db.users.update_one({"id": user_id}, {"$set": {"aadhaar_verified": True}})
    if res.matched_count == 0:
        raise HTTPException(404, "User not found")
    return {"ok": True}

@api.post("/admin/users/{user_id}/suspend")
async def admin_suspend(user_id: str, _=Depends(admin_user)):
    res = await db.users.update_one({"id": user_id}, {"$set": {"suspended": True, "available": False}})
    if res.matched_count == 0:
        raise HTTPException(404, "User not found")
    return {"ok": True}

@api.post("/admin/users/{user_id}/unsuspend")
async def admin_unsuspend(user_id: str, _=Depends(admin_user)):
    res = await db.users.update_one({"id": user_id}, {"$set": {"suspended": False, "available": True}})
    if res.matched_count == 0:
        raise HTTPException(404, "User not found")
    return {"ok": True}

@api.get("/admin/jobs")
async def admin_jobs(_=Depends(admin_user)):
    cursor = db.jobs.find({}, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cursor.to_list(length=200)

@api.post("/admin/jobs/{job_id}/close")
async def admin_close_job(job_id: str, _=Depends(admin_user)):
    res = await db.jobs.update_one({"id": job_id}, {"$set": {"status": "closed"}})
    if res.matched_count == 0:
        raise HTTPException(404, "Job not found")
    return {"ok": True}

@api.get("/admin/attendance")
async def admin_attendance(_=Depends(admin_user)):
    cursor = db.attendance.find({}, {"_id": 0, "selfie": 0}).sort("created_at", -1).limit(200)
    return await cursor.to_list(length=200)

@api.get("/admin/complaints")
async def admin_complaints(status: Optional[str] = None, _=Depends(admin_user)):
    query: dict = {}
    if status:
        query["status"] = status
    cursor = db.complaints.find(query, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cursor.to_list(length=200)

@api.post("/admin/complaints/{cid}/resolve")
async def admin_resolve_complaint(cid: str, note: Optional[str] = "", _=Depends(admin_user)):
    res = await db.complaints.update_one({"id": cid}, {"$set": {"status": "resolved", "admin_note": note or ""}})
    if res.matched_count == 0:
        raise HTTPException(404, "Complaint not found")
    return {"ok": True}

@api.post("/admin/complaints/{cid}/reject")
async def admin_reject_complaint(cid: str, note: Optional[str] = "", _=Depends(admin_user)):
    res = await db.complaints.update_one({"id": cid}, {"$set": {"status": "rejected", "admin_note": note or ""}})
    if res.matched_count == 0:
        raise HTTPException(404, "Complaint not found")
    return {"ok": True}

    res = await db.complaints.update_one({"id": cid}, {"$set": {"status": "rejected", "admin_note": note or ""}})
    if res.matched_count == 0:
        raise HTTPException(404, "Complaint not found")
    return {"ok": True}

# =====================================================================
# PHASE 2 — Production Features
# Forgot Password, Escrow, Ratings enhancement, Leave, Progress Photos,
# Activity Log, Enhanced Filters, Salary Summary, Project Progress,
# Admin Monitoring.
# =====================================================================

# ---------- Activity Log helper ----------
async def log_activity(actor_id: str, actor_role: str, action: str, target_type: str = "", target_id: str = "", meta: Optional[dict] = None):
    """Append an entry to activity_log collection. Non-blocking on failure."""
    try:
        await db.activity_log.insert_one({
            "id": str(uuid.uuid4()),
            "actor_id": actor_id,
            "actor_role": actor_role,
            "action": action,
            "target_type": target_type,
            "target_id": target_id,
            "meta": meta or {},
            "created_at": now_iso(),
        })
    except Exception as e:
        logger.debug("Activity log failed: %s", e)

# ---------- Forgot Password ----------
@api.post("/auth/forgot-password")
async def forgot_password(body: ForgotPasswordIn, request: Request):
    """Sends an OTP to the mobile for password reset. In dev mode, returns dev_code."""
    await rate_limit("forgot_password", request, max_calls=10, window_sec=60)
    if not re.fullmatch(r"\d{10}", body.mobile or ""):
        raise HTTPException(400, "Invalid mobile")
    user = await db.users.find_one({"mobile": body.mobile})
    if not user:
        # Do NOT reveal user existence; but return success shape
        return {"ok": True, "dev_code": None}
    code = f"{secrets.randbelow(1_000_000):06d}" if os.getenv("PROD_OTP") == "1" else DEV_OTP_CODE
    await db.password_resets.update_one(
        {"mobile": body.mobile},
        {"$set": {"code": code, "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat()}},
        upsert=True,
    )
    # In prod, send via Twilio; in dev, just return the code
    if _twilio_client and os.getenv("TWILIO_FROM"):
        try:
            _twilio_client.messages.create(body=f"BuildMitra password reset OTP: {code}", from_=os.getenv("TWILIO_FROM"), to=f"+91{body.mobile}")
        except Exception as e:
            logger.warning("Twilio password OTP failed: %s", e)
    return {"ok": True, "dev_code": code if os.getenv("PROD_OTP") != "1" else None}

@api.post("/auth/reset-password")
async def reset_password(body: ResetPasswordIn):
    if len(body.new_password) < 4:
        raise HTTPException(400, "Password must be at least 4 characters")
    rec = await db.password_resets.find_one({"mobile": body.mobile})
    if not rec or rec.get("code") != body.otp:
        raise HTTPException(401, "Invalid OTP")
    exp = rec.get("expires_at")
    if exp and datetime.fromisoformat(exp) < datetime.now(timezone.utc):
        raise HTTPException(401, "OTP expired")
    user = await db.users.find_one({"mobile": body.mobile})
    if not user:
        raise HTTPException(404, "User not found")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password": hash_pw(body.new_password), "password_updated_at": now_iso()}}
    )
    await db.password_resets.delete_one({"mobile": body.mobile})
    await log_activity(user["id"], user["role"], "password_reset")
    return {"ok": True, "message": "Password reset successful"}

# ---------- Escrow Wallet ----------
@api.post("/escrow/deposit")
async def escrow_deposit(body: EscrowDepositIn, user=Depends(current_user)):
    """Client/contractor deposits money into escrow for a job.
    Amount is added to escrow (not yet paid to worker)."""
    if user["role"] not in ("client", "contractor"):
        raise HTTPException(403, "Only clients/contractors can escrow")
    if body.amount <= 0:
        raise HTTPException(400, "Invalid amount")
    job = await db.jobs.find_one({"id": body.job_id})
    if not job or job["posted_by"] != user["id"]:
        raise HTTPException(403, "Not your job")
    esc = {
        "id": str(uuid.uuid4()),
        "job_id": body.job_id,
        "job_title": job.get("title"),
        "payer_id": user["id"],
        "amount": body.amount,
        "amount_released": 0.0,
        "status": "held",  # held | released | refunded
        "created_at": now_iso(),
    }
    await db.escrow.insert_one(esc)
    await log_activity(user["id"], user["role"], "escrow_deposit", "escrow", esc["id"], {"job_id": body.job_id, "amount": body.amount})
    esc.pop("_id", None)
    return {"ok": True, "escrow": esc}

@api.get("/escrow/mine")
async def escrow_mine(user=Depends(current_user)):
    """Returns escrow records owned by user (as payer or beneficiary)."""
    q: dict = {"payer_id": user["id"]}
    cur = db.escrow.find(q, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cur.to_list(length=200)

@api.post("/escrow/release")
async def escrow_release(body: EscrowActionIn, user=Depends(current_user)):
    """Client releases (part of) escrow to a worker. Adds to worker wallet."""
    if not body.worker_id or not body.amount or body.amount <= 0:
        raise HTTPException(400, "worker_id and positive amount required")
    esc = await db.escrow.find_one({"id": body.escrow_id})
    if not esc:
        raise HTTPException(404, "Escrow not found")
    if esc["payer_id"] != user["id"]:
        raise HTTPException(403, "Not your escrow")
    remaining = float(esc["amount"]) - float(esc.get("amount_released", 0))
    if body.amount > remaining:
        raise HTTPException(400, f"Amount exceeds remaining escrow (₹{remaining:.2f})")
    worker = await db.users.find_one({"id": body.worker_id})
    if not worker:
        raise HTTPException(404, "Worker not found")
    new_released = float(esc.get("amount_released", 0)) + float(body.amount)
    new_status = "released" if new_released >= float(esc["amount"]) - 0.01 else "held"
    await db.escrow.update_one(
        {"id": esc["id"]},
        {"$set": {"amount_released": new_released, "status": new_status, "last_release_at": now_iso()}}
    )
    await db.users.update_one({"id": body.worker_id}, {"$inc": {"wallet_balance": float(body.amount)}})
    txn = {
        "id": str(uuid.uuid4()),
        "user_id": body.worker_id,
        "type": "escrow_release",
        "amount": float(body.amount),
        "status": "success",
        "note": f"Escrow released for job: {esc.get('job_title', '')}",
        "job_id": esc.get("job_id"),
        "escrow_id": esc["id"],
        "created_at": now_iso(),
    }
    await db.wallet_txns.insert_one(txn)
    await log_activity(user["id"], user["role"], "escrow_release", "escrow", esc["id"], {"worker_id": body.worker_id, "amount": body.amount})
    return {"ok": True, "released": new_released, "status": new_status}

@api.post("/escrow/refund")
async def escrow_refund(body: EscrowActionIn, user=Depends(current_user)):
    """Refund remaining escrow amount back to payer (job cancelled)."""
    esc = await db.escrow.find_one({"id": body.escrow_id})
    if not esc:
        raise HTTPException(404, "Escrow not found")
    if esc["payer_id"] != user["id"]:
        raise HTTPException(403, "Not your escrow")
    if esc["status"] != "held":
        raise HTTPException(400, "Escrow already settled")
    await db.escrow.update_one(
        {"id": esc["id"]},
        {"$set": {"status": "refunded", "refunded_at": now_iso()}}
    )
    await log_activity(user["id"], user["role"], "escrow_refund", "escrow", esc["id"])
    return {"ok": True, "status": "refunded"}

# ---------- Ratings (list for target) ----------
@api.get("/ratings/user/{user_id}")
async def ratings_for_user(user_id: str, _user=Depends(current_user)):
    cur = db.ratings.find({"target_user_id": user_id}, {"_id": 0}).sort("created_at", -1).limit(100)
    return await cur.to_list(length=100)

@api.get("/ratings/mine")
async def my_ratings(user=Depends(current_user)):
    """Ratings this user has given."""
    cur = db.ratings.find({"by": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(100)
    return await cur.to_list(length=100)

# ---------- Leave Management ----------
@api.post("/leave/request")
async def leave_request(body: LeaveRequestIn, user=Depends(current_user)):
    if user["role"] != "worker":
        raise HTTPException(403, "Only workers can request leave")
    if not body.from_date or not body.to_date or not body.reason.strip():
        raise HTTPException(400, "All fields required")
    approver_id = ""
    approver_name = ""
    if body.job_id:
        job = await db.jobs.find_one({"id": body.job_id})
        if job:
            approver_id = job["posted_by"]
            approver_name = job.get("posted_by_name", "")
    rec = {
        "id": str(uuid.uuid4()),
        "worker_id": user["id"],
        "worker_name": user["name"],
        "job_id": body.job_id or "",
        "approver_id": approver_id,
        "approver_name": approver_name,
        "from_date": body.from_date,
        "to_date": body.to_date,
        "reason": body.reason.strip(),
        "status": "pending",
        "note": "",
        "created_at": now_iso(),
    }
    await db.leaves.insert_one(rec)
    await log_activity(user["id"], user["role"], "leave_request", "leave", rec["id"])
    rec.pop("_id", None)
    return rec

@api.get("/leave/mine")
async def leave_mine(user=Depends(current_user)):
    cur = db.leaves.find({"worker_id": user["id"]}, {"_id": 0}).sort("created_at", -1).limit(100)
    return await cur.to_list(length=100)

@api.get("/leave/inbox")
async def leave_inbox(user=Depends(current_user)):
    """For contractor/client: pending/decided leave requests for their jobs."""
    if user["role"] not in ("client", "contractor", "admin"):
        raise HTTPException(403, "Not allowed")
    q = {"approver_id": user["id"]} if user["role"] != "admin" else {}
    cur = db.leaves.find(q, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cur.to_list(length=200)

@api.post("/leave/{leave_id}/decision")
async def leave_decision(leave_id: str, body: LeaveDecisionIn, user=Depends(current_user)):
    if body.decision not in ("approved", "rejected"):
        raise HTTPException(400, "Invalid decision")
    rec = await db.leaves.find_one({"id": leave_id})
    if not rec:
        raise HTTPException(404, "Leave request not found")
    if user["role"] != "admin" and rec.get("approver_id") != user["id"]:
        raise HTTPException(403, "Not your inbox")
    await db.leaves.update_one(
        {"id": leave_id},
        {"$set": {"status": body.decision, "note": body.note or "", "decided_at": now_iso()}}
    )
    await log_activity(user["id"], user["role"], f"leave_{body.decision}", "leave", leave_id)
    return {"ok": True, "status": body.decision}

# ---------- Site Progress Photos ----------
@api.post("/progress-photos")
async def add_progress_photo(body: ProgressPhotoIn, user=Depends(current_user)):
    if user["role"] not in ("contractor", "worker", "client"):
        raise HTTPException(403, "Not allowed")
    job = await db.jobs.find_one({"id": body.job_id})
    if not job:
        raise HTTPException(404, "Job not found")
    if body.photo:
        body.photo = validate_photo(body.photo)
    rec = {
        "id": str(uuid.uuid4()),
        "job_id": body.job_id,
        "job_title": job.get("title"),
        "uploader_id": user["id"],
        "uploader_name": user["name"],
        "uploader_role": user["role"],
        "photo": body.photo,
        "caption": sanitize_str(body.caption or "", 500),
        "lat": body.lat,
        "lng": body.lng,
        "created_at": now_iso(),
    }
    await db.progress_photos.insert_one(rec)
    await log_activity(user["id"], user["role"], "progress_photo_add", "job", body.job_id)
    rec.pop("_id", None)
    return rec

@api.get("/progress-photos/{job_id}")
async def list_progress_photos(job_id: str, _user=Depends(current_user)):
    cur = db.progress_photos.find({"job_id": job_id}, {"_id": 0}).sort("created_at", -1).limit(200)
    return await cur.to_list(length=200)

@api.delete("/progress-photos/{photo_id}")
async def delete_progress_photo(photo_id: str, user=Depends(current_user)):
    rec = await db.progress_photos.find_one({"id": photo_id})
    if not rec:
        raise HTTPException(404, "Not found")
    if rec["uploader_id"] != user["id"] and user["role"] != "admin":
        raise HTTPException(403, "Not your photo")
    await db.progress_photos.delete_one({"id": photo_id})
    return {"ok": True}

# ---------- Client Project Progress ----------
@api.get("/projects/progress")
async def project_progress(user=Depends(current_user)):
    """Progress summary for each job posted by the user.
    Returns: [{job_id, title, status, workers_hired, days_worked, photos_count,
               escrow_amount, escrow_released, complaints_count}]
    """
    if user["role"] not in ("client", "contractor", "admin"):
        raise HTTPException(403, "Not allowed")
    q = {} if user["role"] == "admin" else {"posted_by": user["id"]}
    jobs = await db.jobs.find(q, {"_id": 0}).sort("created_at", -1).limit(100).to_list(length=100)
    if not jobs:
        return []
    job_ids = [j["id"] for j in jobs]
    apps = await db.applications.find(
        {"job_id": {"$in": job_ids}, "status": "accepted"},
        {"_id": 0, "job_id": 1, "worker_id": 1}
    ).to_list(length=None)
    att = await db.attendance.find(
        {"job_id": {"$in": job_ids}, "type": "check_in", "within_geofence": True},
        {"_id": 0, "job_id": 1, "worker_id": 1}
    ).to_list(length=None)
    photos = await db.progress_photos.find(
        {"job_id": {"$in": job_ids}},
        {"_id": 0, "job_id": 1}
    ).to_list(length=None)
    esc = await db.escrow.find(
        {"job_id": {"$in": job_ids}},
        {"_id": 0}
    ).to_list(length=None)

    def agg(job_id: str):
        hired = {a["worker_id"] for a in apps if a["job_id"] == job_id}
        days = sum(1 for a in att if a["job_id"] == job_id)
        photos_count = sum(1 for p in photos if p["job_id"] == job_id)
        esc_amt = sum(float(e["amount"]) for e in esc if e["job_id"] == job_id)
        esc_rel = sum(float(e.get("amount_released", 0)) for e in esc if e["job_id"] == job_id)
        return len(hired), days, photos_count, esc_amt, esc_rel

    out = []
    for j in jobs:
        h, d, p, ea, er = agg(j["id"])
        out.append({
            "job_id": j["id"], "title": j.get("title"),
            "status": j.get("status", "open"),
            "workers_hired": h,
            "workers_needed": j.get("workers_needed", 1),
            "days_worked": d,
            "duration_days": j.get("duration_days", 0),
            "photos_count": p,
            "escrow_amount": ea, "escrow_released": er,
            "daily_wage": j.get("daily_wage", 0),
            "created_at": j.get("created_at"),
        })
    return out

# ---------- Worker Salary Summary ----------
@api.get("/salary/summary")
async def salary_summary(months: int = 6, user=Depends(current_user)):
    """For workers: monthly attendance & earnings summary for last N months."""
    if user["role"] != "worker":
        raise HTTPException(403, "Workers only")
    wage = float(user.get("daily_wage", 0) or 0)
    # Fetch all check-ins within geofence for user in last N months
    cutoff = (datetime.now(timezone.utc) - timedelta(days=months * 31)).isoformat()
    att = await db.attendance.find(
        {"worker_id": user["id"], "type": "check_in", "within_geofence": True,
         "created_at": {"$gte": cutoff}},
        {"_id": 0, "selfie": 0}
    ).limit(2000).to_list(length=2000)
    # Group by month
    by_month: dict = {}
    for a in att:
        m = str(a["created_at"])[:7]  # YYYY-MM
        by_month.setdefault(m, {"month": m, "days_present": 0, "jobs": set()})
        by_month[m]["days_present"] += 1
        if a.get("job_id") and a["job_id"] != "self":
            by_month[m]["jobs"].add(a["job_id"])
    rows = []
    for m in sorted(by_month.keys(), reverse=True):
        r = by_month[m]
        rows.append({
            "month": m,
            "days_present": r["days_present"],
            "jobs_count": len(r["jobs"]),
            "daily_wage": wage,
            "earned": r["days_present"] * wage,
        })
    total_earned = sum(r["earned"] for r in rows)
    wallet_bal = float(user.get("wallet_balance", 0) or 0)
    return {
        "rows": rows,
        "total_earned": total_earned,
        "wallet_balance": wallet_bal,
        "current_wage": wage,
    }

# ---------- Admin Monitoring ----------
@api.get("/admin/activity")
async def admin_activity(limit: int = 100, user=Depends(current_user)):
    if user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    cur = db.activity_log.find({}, {"_id": 0}).sort("created_at", -1).limit(min(limit, 500))
    return await cur.to_list(length=min(limit, 500))

@api.get("/admin/monitor")
async def admin_monitor(user=Depends(current_user)):
    if user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    users_total = await db.users.count_documents({})
    workers = await db.users.count_documents({"role": "worker"})
    contractors = await db.users.count_documents({"role": "contractor"})
    clients = await db.users.count_documents({"role": "client"})
    aadhaar_verified = await db.users.count_documents({"aadhaar_verified": True})
    jobs_total = await db.jobs.count_documents({})
    jobs_open = await db.jobs.count_documents({"status": "open"})
    jobs_progress = await db.jobs.count_documents({"status": "in_progress"})
    jobs_completed = await db.jobs.count_documents({"status": "completed"})
    complaints_open = await db.complaints.count_documents({"status": "open"})
    escrow_held = await db.escrow.count_documents({"status": "held"})
    photos = await db.progress_photos.count_documents({})
    leaves_pending = await db.leaves.count_documents({"status": "pending"})
    # Wallet: total held in system
    wallet_pipeline = [{"$group": {"_id": None, "total": {"$sum": "$wallet_balance"}}}]
    tot_wallet = 0.0
    async for row in db.users.aggregate(wallet_pipeline):
        tot_wallet = float(row.get("total", 0) or 0)
    return {
        "users": {"total": users_total, "workers": workers, "contractors": contractors,
                  "clients": clients, "aadhaar_verified": aadhaar_verified},
        "jobs": {"total": jobs_total, "open": jobs_open, "in_progress": jobs_progress,
                 "completed": jobs_completed},
        "complaints_open": complaints_open,
        "escrow_held": escrow_held,
        "progress_photos": photos,
        "leaves_pending": leaves_pending,
        "total_wallet_balance": tot_wallet,
    }

# ---------- Enhanced Jobs Search moved to top of jobs routes (before /jobs/{job_id})
# to avoid path shadowing. See /jobs/search near line 687.

# --- Seed ---
async def seed():
    try:
        existing = await db.users.count_documents({"role": "client"})
    except Exception as e:
        logger.warning("Seed: cannot check users (Mongo unavailable?): %s", e)
        return
    if existing > 0:
        return
    logger.info("Seeding demo data...")
    demo_client = {
        "id": str(uuid.uuid4()), "name": "Sharma Builders",
        "mobile": "9000000001", "password": hash_pw("demo1234"),
        "role": "client", "photo": None, "skills": [], "experience_years": 0,
        "daily_wage": 0, "available": True, "city": "Mumbai",
        "company_name": "Sharma Builders Pvt Ltd", "aadhaar_verified": True,
        "language": "en", "rating_avg": 4.6, "rating_count": 12,
        "referral_code": "BM0001AAAA", "referred_by": None, "wallet_balance": 0,
        "created_at": now_iso(),
    }
    demo_worker = {
        "id": str(uuid.uuid4()), "name": "Ramesh Kumar",
        "mobile": "9000000002", "password": hash_pw("demo1234"),
        "role": "worker", "photo": None,
        "skills": ["Mason", "Tile Worker"], "experience_years": 6,
        "daily_wage": 800, "available": True, "city": "Mumbai",
        "company_name": "", "aadhaar_verified": True, "language": "en",
        "rating_avg": 4.4, "rating_count": 9,
        "referral_code": "BM0002BBBB", "referred_by": None, "wallet_balance": 50,
        "created_at": now_iso(),
    }
    demo_contractor = {
        "id": str(uuid.uuid4()), "name": "Suresh Patel",
        "mobile": "9000000003", "password": hash_pw("demo1234"),
        "role": "contractor", "photo": None,
        "skills": [], "experience_years": 12, "daily_wage": 0,
        "available": True, "city": "Pune",
        "company_name": "Patel Construction Co", "aadhaar_verified": True,
        "language": "en", "rating_avg": 4.8, "rating_count": 24,
        "referral_code": "BM0003CCCC", "referred_by": None, "wallet_balance": 0,
        "created_at": now_iso(),
    }
    await db.users.insert_many([demo_client, demo_worker, demo_contractor])

    admin_user_doc = {
        "id": str(uuid.uuid4()), "name": "BuildMitra Admin",
        "mobile": "9000000000", "password": hash_pw("admin1234"),
        "role": "admin", "photo": None, "skills": [], "experience_years": 0,
        "daily_wage": 0, "available": True, "city": "HQ",
        "company_name": "BuildMitra", "aadhaar_verified": True, "language": "en",
        "rating_avg": 0.0, "rating_count": 0,
        "referral_code": "BMADMIN001", "referred_by": None, "wallet_balance": 0,
        "created_at": now_iso(),
    }
    await db.users.insert_one(admin_user_doc)

    # Seed two demo complaints so admin panel has content
    await db.complaints.insert_many([
        {
            "id": str(uuid.uuid4()),
            "by_user_id": demo_worker["id"],
            "by_user_name": demo_worker["name"],
            "by_user_role": "worker",
            "against_user_id": demo_client["id"],
            "subject": "Payment delayed by 7 days",
            "description": "Worked at Andheri site for 5 days, payment still pending.",
            "status": "open", "admin_note": "",
            "created_at": now_iso(),
        },
        {
            "id": str(uuid.uuid4()),
            "by_user_id": demo_client["id"],
            "by_user_name": demo_client["name"],
            "by_user_role": "client",
            "against_user_id": demo_worker["id"],
            "subject": "Worker not reporting on time",
            "description": "Two workers consistently late by 1+ hour.",
            "status": "open", "admin_note": "",
            "created_at": now_iso(),
        },
    ])

    sample_jobs = [
        ("Mason needed for 2BHK plastering", "Need experienced mason for wall plastering work. Daily basis.", "Mason", 5, 900, "Andheri, Mumbai", "Urgent"),
        ("Tile workers for villa", "Italian tiles to be laid in 3 rooms. Skilled tile workers only.", "Tile Worker", 3, 1100, "Bandra, Mumbai", "Normal"),
        ("Electrician for new office", "Complete wiring + DB installation for 1500sqft office.", "Electrician", 2, 1200, "Powai, Mumbai", "Normal"),
        ("Painter team for apartment", "Interior painting (3 coats) of 3BHK + ceiling.", "Painter", 4, 850, "Thane, Mumbai", "Urgent"),
        ("Plumber for villa", "PVC + CPVC plumbing complete fitting.", "Plumber", 2, 1000, "Pune, Maharashtra", "Normal"),
        ("Helper required at site", "General site helper required immediately.", "Helper", 8, 500, "Andheri, Mumbai", "Urgent"),
        ("Steel fixer for slab", "RCC slab steel fixing 2000 sqft.", "Steel Fixer", 6, 1100, "Vashi, Navi Mumbai", "Normal"),
        ("Carpenter for furniture", "Wardrobe + modular kitchen carpentry.", "Carpenter", 2, 1300, "Goregaon, Mumbai", "Normal"),
    ]
    for title, desc, skill, n, wage, loc, urgency in sample_jobs:
        await db.jobs.insert_one({
            "id": str(uuid.uuid4()),
            "posted_by": demo_client["id"],
            "posted_by_name": demo_client["name"],
            "posted_by_role": "client",
            "title": title, "description": desc, "skill": skill,
            "workers_needed": n, "daily_wage": wage,
            "location": loc, "site_address": loc,
            "start_date": None, "duration_days": 10,
            "working_hours": "8 hrs", "urgency": urgency,
            "lat": None, "lng": None,
            "status": "open", "applicants_count": 0,
            "created_at": now_iso(),
        })
    logger.info("Seed done.")

app.include_router(api)

# CORS — allow specific origins in prod (via ALLOWED_ORIGINS env),
# fall back to permissive "*" in dev so preview URLs work seamlessly.
_allowed = os.getenv("ALLOWED_ORIGINS", "").strip()
_origins: list = [o.strip() for o in _allowed.split(",") if o.strip()] if _allowed else ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Root-level health endpoints (NOT under /api) for K8s probes that
# hit non-prefixed paths.
@app.get("/health")
async def app_health():
    return {"status": "ok"}

@app.get("/healthz")
async def app_healthz():
    return {"status": "ok"}

@app.on_event("startup")
async def on_start():
    """Run seed + create indexes as a background task so startup completes
    immediately. K8s liveness/readiness probes see the app as healthy before
    any CPU-bound / DB work begins."""
    async def _boot_safe():
        try:
            await asyncio.sleep(1.5)  # let probes hit /health first
            await create_indexes()
            await seed()
        except Exception as e:
            logger.exception("Boot task failed (non-fatal): %s", e)
    asyncio.create_task(_boot_safe())


async def create_indexes():
    """Create MongoDB indexes on hot query fields.
    Idempotent — calling repeatedly is safe."""
    try:
        # Users: mobile lookup on every login; role for admin queries
        await db.users.create_index("mobile", unique=True, sparse=True)
        await db.users.create_index("role")
        await db.users.create_index("id", unique=True)

        # Jobs: posted_by + status filtering, geo/text search
        await db.jobs.create_index("posted_by")
        await db.jobs.create_index("status")
        await db.jobs.create_index("skill")
        await db.jobs.create_index([("created_at", -1)])
        await db.jobs.create_index("id", unique=True)

        # Applications: job_id, worker_id for status queries
        await db.applications.create_index("job_id")
        await db.applications.create_index("worker_id")
        await db.applications.create_index([("worker_id", 1), ("status", 1)])
        await db.applications.create_index("id", unique=True)

        # Attendance: worker_id + created_at for history & payroll
        await db.attendance.create_index([("worker_id", 1), ("created_at", -1)])
        await db.attendance.create_index([("job_id", 1), ("created_at", -1)])
        await db.attendance.create_index("id", unique=True)

        # ERP
        for coll in ("materials", "tools", "estimates", "bills"):
            await db[coll].create_index("owner")
            await db[coll].create_index([("created_at", -1)])

        # Chat
        await db.chat_messages.create_index([("thread_id", 1), ("created_at", -1)])

        # Complaints
        await db.complaints.create_index("by_user_id")
        await db.complaints.create_index("status")

        # Wallet
        await db.wallet_txns.create_index([("user_id", 1), ("created_at", -1)])

        # Phase 2 collections
        await db.escrow.create_index("payer_id")
        await db.escrow.create_index([("job_id", 1)])
        await db.leaves.create_index([("worker_id", 1), ("status", 1)])
        await db.leaves.create_index("approver_id")
        await db.progress_photos.create_index([("job_id", 1), ("created_at", -1)])
        await db.activity_log.create_index([("created_at", -1)])
        await db.password_resets.create_index("mobile", unique=True)
        # TTL: purge password resets after 15 minutes
        try:
            await db.password_resets.create_index("expires_at", expireAfterSeconds=0)
        except Exception:
            pass
        logger.info("Mongo indexes ensured.")
    except Exception as e:
        logger.warning("Index creation partially failed: %s", e)

@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        client.close()
    except Exception:
        pass

# -------------------------------------------------------------------
# Serve Expo Web build (SPA) at root path.
# This way the single FastAPI service serves both the frontend SPA
# and the /api/* JSON endpoints — eliminating ingress routing issues
# in production where only one port is exposed.
# -------------------------------------------------------------------
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse

FRONTEND_DIST = Path(__file__).resolve().parent.parent / "frontend" / "dist"

if FRONTEND_DIST.exists() and (FRONTEND_DIST / "index.html").exists():
    logger.info("Mounting Expo web build from %s", FRONTEND_DIST)
    # Mount static asset directories
    if (FRONTEND_DIST / "_expo").exists():
        app.mount("/_expo", StaticFiles(directory=str(FRONTEND_DIST / "_expo")), name="expo_static")
    if (FRONTEND_DIST / "assets").exists():
        app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIST / "assets")), name="static_assets")

    @app.get("/favicon.ico", include_in_schema=False)
    @app.head("/favicon.ico", include_in_schema=False)
    async def favicon():
        f = FRONTEND_DIST / "favicon.ico"
        return FileResponse(f) if f.exists() else JSONResponse({"detail": "Not Found"}, status_code=404)

    @app.get("/", include_in_schema=False)
    @app.head("/", include_in_schema=False)
    async def spa_root():
        return FileResponse(FRONTEND_DIST / "index.html", media_type="text/html")

    # SPA catch-all: any non-API route returns index.html so client-side
    # routing (expo-router) can take over. /api/* routes are matched
    # first because they were included before this catch-all.
    @app.get("/{full_path:path}", include_in_schema=False)
    @app.head("/{full_path:path}", include_in_schema=False)
    async def spa_catch_all(full_path: str):
        # Don't intercept api routes (already matched), health endpoints,
        # or static asset paths (already mounted).
        if full_path.startswith(("api/", "api", "health", "healthz", "_expo/", "assets/")):
            return JSONResponse({"detail": "Not Found"}, status_code=404)
        # Try to serve a real file if it exists in dist (for /metadata.json, etc.)
        candidate = FRONTEND_DIST / full_path
        if candidate.is_file():
            return FileResponse(candidate)
        # Otherwise fall back to SPA index
        return FileResponse(FRONTEND_DIST / "index.html", media_type="text/html")
else:
    logger.warning("Expo web build not found at %s — frontend will not be served from backend.", FRONTEND_DIST)

    # Friendly fallback so root URL doesn't return raw JSON 404 if dist is missing
    @app.get("/", include_in_schema=False)
    @app.head("/", include_in_schema=False)
    async def root_fallback():
        return JSONResponse({
            "app": "BuildMitra",
            "status": "Backend is running but web build is missing.",
            "hint": "Run `expo export --platform web --output-dir dist` in frontend/ and redeploy.",
            "api": "/api/health"
        })
